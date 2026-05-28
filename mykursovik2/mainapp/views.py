import os
import json as _json
import chardet
import tempfile
from decimal import Decimal as _Decimal
from io import StringIO

from django.apps import apps
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.files.uploadedfile import UploadedFile
from django.core.management import call_command
from django.core.paginator import Paginator, EmptyPage
from django.db import transaction, connection
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.views.decorators.http import require_POST

from .models import (
    CarMake, CarModel, Client, ClientCar, Order, Service, ServiceType
)
from .forms import (
    ClientForm, ClientCarForm, CarMakeForm, CarModelForm,
    ServiceTypeForm, ServiceForm, ClientSelectionForm, OrderForm
)
from .validators import normalize_plate, normalize_plate_search

@login_required
def clients_list(request):
    # Регистрируем кастомную SQL-функцию
    register_custom_functions()

    clients = Client.objects.all()

    # Получаем параметры запроса
    sort = request.GET.get('sort', 'id')
    direction = request.GET.get('direction', 'desc')

    # Обработка сортировки
    if sort in ['id', 'phone']:
        if direction == 'desc':
            clients = clients.order_by(f'-{sort}')
        else:
            clients = clients.order_by(sort)
    elif sort == 'fio':
        # Сортировка по ФИО с использованием custom_lower
        if direction == 'desc':
            clients = clients.extra(select={'fio_lower': 'custom_lower(fio)'}).order_by('-fio_lower')
        else:
            clients = clients.extra(select={'fio_lower': 'custom_lower(fio)'}).order_by('fio_lower')

    # Поиск
    search = request.GET.get('search', '')
    column = request.GET.get('column', 'fio')

    if search:
        search_lower = search.lower()
        if column == 'id':
            clients = clients.filter(id__contains=search)
        elif column == 'fio':
            clients = clients.extra(
                select={'fio_lower': 'custom_lower(fio)'},
                where=['custom_lower(fio) LIKE %s'],
                params=[f'%{search_lower}%']
            )
        elif column == 'phone':
            clients = clients.extra(
                select={'phone_lower': 'custom_lower(phone)'},
                where=['custom_lower(phone) LIKE %s'],
                params=[f'%{search_lower}%']
            )

    # Пагинация
    per_page = request.GET.get('per_page', 5)
    try:
        per_page = int(per_page)
        if per_page not in [5, 10, 20]:
            per_page = 5
    except ValueError:
        per_page = 5

    paginator = Paginator(clients, per_page)
    page_number = request.GET.get('page', 1)
    try:
        page_obj = paginator.get_page(page_number)
    except EmptyPage:
        page_obj = paginator.page(1)

    return render(request, 'clients/clients_list.html', {
        'page_obj': page_obj,
        'per_page': per_page,
    })


# Создание клиента
@login_required
def client_create(request):
    if request.method == 'POST':
        form = ClientForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Клиент успешно создан!')
            return redirect('clients_list')
    else:
        form = ClientForm()
    return render(request, 'clients/client_create.html', {'form': form, 'cancel_url': reverse('clients_list')})


import logging

logger = logging.getLogger(__name__)


# Кастомная функция для приведения к нижнему регистру
def custom_lower(text):
    if text is None:
        return None
    return text.lower()  # Python корректно обрабатывает кириллицу


# Регистрация функции в SQLite
def register_custom_functions():
    with connection.cursor() as cursor:
        connection.connection.create_function("custom_lower", 1, custom_lower)
        logger.debug("Custom lower function registered in SQLite")


@login_required
def client_detail(request, pk):
    # Регистрируем функцию при каждом запросе
    register_custom_functions()

    client = get_object_or_404(Client, pk=pk)
    client_cars = ClientCar.objects.filter(client=client)

    # Сортировка
    sort = request.GET.get('sort', 'model__name')
    direction = request.GET.get('direction', 'asc')
    valid_sort_fields = ['make__name', 'model__name', 'license_plate']
    if sort in valid_sort_fields:
        if direction == 'desc':
            client_cars = client_cars.order_by(f'-{sort}')
        else:
            client_cars = client_cars.order_by(sort)

    # Поиск
    search = request.GET.get('search', '')
    column = request.GET.get('column', 'model')

    if search:
        if column == 'make':
            client_cars = client_cars.filter(make__name__contains=search)
        elif column == 'model':
            client_cars = client_cars.filter(model__name__contains=search)
        elif column == 'license_plate':
            from django.db.models import Q
            variants = normalize_plate_search(search)
            q = Q()
            for v in variants:
                q |= Q(license_plate__icontains=v)
            client_cars = client_cars.filter(q)

    # Пагинация
    per_page = request.GET.get('per_page', 5)
    try:
        per_page = int(per_page)
        if per_page not in [5, 10, 20]:
            per_page = 5
    except ValueError:
        per_page = 5

    paginator = Paginator(client_cars, per_page)
    page_number = request.GET.get('page', 1)
    try:
        page_obj = paginator.get_page(page_number)
    except EmptyPage:
        page_obj = paginator.page(1)

    return render(request, 'clients/client_detail.html', {
        'client': client,
        'page_obj': page_obj,
        'per_page': per_page,
        'cancel_url': reverse('clients_list')
    })

# Редактирование клиента
def client_update(request, pk):
    client = get_object_or_404(Client, pk=pk)
    if request.method == 'POST':
        form = ClientForm(request.POST, instance=client)
        if form.is_valid():
            form.save()
            messages.success(request, 'Клиент успешно обновлён!')
            return redirect('clients_list')
    else:
        form = ClientForm(instance=client)
    return render(request, 'clients/client_update.html', {'form': form, 'cancel_url': reverse('clients_list')})


@require_POST
def check_vin_uniqueness(request):
    vin = request.POST.get('vin', '').strip().upper()
    exists = ClientCar.objects.filter(vin=vin).exists()
    return JsonResponse({'exists': exists})


@require_POST
def check_license_plate_uniqueness(request):
    license_plate = request.POST.get('license_plate', '').strip()
    country = request.POST.get('plate_country', 'RU') or 'RU'
    license_plate = normalize_plate(license_plate, country)
    exists = ClientCar.objects.filter(license_plate=license_plate).exists()
    return JsonResponse({'exists': exists})


@require_POST
def check_service_type_uniqueness(request):
    name = request.POST.get("name", "").strip().lower()
    service_type_id = request.POST.get("id", None)
    queryset = ServiceType.objects.all()
    if service_type_id:
        queryset = queryset.exclude(id=service_type_id)
    exists = any(st.name.lower() == name for st in queryset)
    return JsonResponse({"exists": exists})

@require_POST
def check_service_uniqueness(request):
    name = request.POST.get("name", "").strip().lower()
    service_id = request.POST.get("id", None)
    queryset = Service.objects.all()
    if service_id:
        queryset = queryset.exclude(id=service_id)
    exists = any(s.name.lower() == name for s in queryset)
    return JsonResponse({"exists": exists})

@require_POST
def check_car_make_uniqueness(request):
    name = request.POST.get("name", "").strip().lower()
    car_make_id = request.POST.get("id", None)
    queryset = CarMake.objects.all()
    if car_make_id:
        queryset = queryset.exclude(id=car_make_id)
    exists = any(cm.name.lower() == name for cm in queryset)
    return JsonResponse({"exists": exists})

@require_POST
def check_car_model_uniqueness(request):
    name = request.POST.get("name", "").strip().lower()
    car_model_id = request.POST.get("id", None)
    queryset = CarModel.objects.all()
    if car_model_id:
        queryset = queryset.exclude(id=car_model_id)
    exists = any(cm.name.lower() == name for cm in queryset)
    return JsonResponse({"exists": exists})


@require_POST
def check_client_phone_uniqueness(request):
    phone = request.POST.get("phone", "").strip()
    client_id = request.POST.get("id", None)
    queryset = Client.objects.all()
    if client_id:
        queryset = queryset.exclude(id=client_id)
    exists = queryset.filter(phone=phone).exists()
    return JsonResponse({"exists": exists})

# Удаление клиента
def client_delete(request, pk):
    client = get_object_or_404(Client, pk=pk)
    if request.method == 'POST':
        client.delete()
        messages.success(request, 'Клиент успешно удалён!')
        return redirect('clients_list')
    return render(request, 'confirm_delete.html', {
        'object': client,
        'deleted_object_type': 'Client',
        'cancel_url': reverse('clients_list')
    })


# Просмотр списка автомобилей клиента с пагинацией
def client_cars_list(request):
    client_cars = ClientCar.objects.all()
    paginator = Paginator(client_cars, 5)  # Показывать 5 автомобилей на страницу
    page_number = request.GET.get('page', 1)
    try:
        page_obj = paginator.get_page(page_number)
    except EmptyPage:
        page_obj = paginator.page(1)  # Если страница не найдена, показываем первую страницу
    return render(request, 'clients/client_cars_list.html', {'page_obj': page_obj, 'per_page': 5})


# Создание автомобиля для клиента (одна форма с динамическим выбором модели)
def client_car_create(request, pk):
    client = get_object_or_404(Client, pk=pk)
    if request.method == 'POST':
        make_id = request.POST.get('make')
        form = ClientCarForm(request.POST, make_id=make_id if make_id else None)
        if form.is_valid():
            client_car = form.save(commit=False)
            client_car.client = client  # Устанавливаем client вручную (для надежности)
            client_car.save()
            messages.success(request, f'Автомобиль успешно создан для клиента {client.fio}!')
            return redirect('client_detail', pk=pk)
        # else:
        #     print("Форма не валидна:", form.errors)  # Отладка
    else:
        form = ClientCarForm(initial={'client': client})  # Задаем начальное значение
    return render(request, 'clients/client_car_create.html', {
        'form': form,
        'client': client,
        'cancel_url': reverse('client_detail', kwargs={'pk': pk})
    })


# # # Получение списка моделей по ID марки (AJAX)
def get_models(request):
    make_id = request.GET.get('make_id')
    if make_id:
        models = CarModel.objects.filter(make_id=make_id).values('id', 'name')
        return JsonResponse(list(models), safe=False)
    return JsonResponse([], safe=False)


def get_all_makes(request):
    page = int(request.GET.get('page', 1))
    per_page = int(request.GET.get('per_page', 10))
    makes = CarMake.objects.all().order_by('name')
    paginator = Paginator(makes, per_page)
    page_obj = paginator.get_page(page)
    data = {
        'makes': [{'id': make.id, 'name': make.name} for make in page_obj],
        'total_pages': paginator.num_pages,
    }
    return JsonResponse(data)


def get_services(request):
    service_type_id = request.GET.get('service_type_id')
    if service_type_id:
        services = Service.objects.filter(service_type_id=service_type_id).values('id', 'name')
        return JsonResponse(list(services), safe=False)
    return JsonResponse([], safe=False)


@login_required
def get_all_clients(request):
    # Регистрируем кастомную функцию
    register_custom_functions()

    clients = Client.objects.all()

    # Поиск только по fio
    search = request.GET.get('search', '')
    if search:
        search_lower = search.lower()
        clients = clients.extra(
            select={'fio_lower': 'custom_lower(fio)'},
            where=['custom_lower(fio) LIKE %s'],
            params=[f'%{search_lower}%']
        )

    # Пагинация
    page = int(request.GET.get('page', 1))
    per_page = int(request.GET.get('per_page', 10))
    total_clients = clients.count()
    total_pages = (total_clients + per_page - 1) // per_page

    start = (page - 1) * per_page
    end = start + per_page
    paginated_clients = clients[start:end].values('id', 'fio', 'phone')

    return JsonResponse({
        'clients': list(paginated_clients),
        'total_pages': total_pages
    }, safe=False)

def get_client_cars(request):
    client_id = request.GET.get('client_id')
    if client_id:
        client_cars = ClientCar.objects.filter(client_id=client_id).values('id', 'make__name', 'model__name',
                                                                           'license_plate')
        return JsonResponse(list(client_cars), safe=False)
    return JsonResponse([], safe=False)


def get_service_types(request):
    types = ServiceType.objects.all().values('id', 'name')
    return JsonResponse(list(types), safe=False)


def client_car_detail(request, pk, car_pk):
    client_car = get_object_or_404(ClientCar, pk=car_pk)
    return render(request, 'clients/client_car_detail.html', {
        'client_car': client_car,
        'cancel_url': reverse('client_detail', kwargs={'pk': client_car.client.pk})
    })


def client_car_update(request, pk, car_pk):
    client_car = get_object_or_404(ClientCar, pk=car_pk)
    client = client_car.client  # Получаем клиента из объекта автомобиля
    if request.method == 'POST':
        make_id = request.POST.get('make')
        form = ClientCarForm(request.POST, instance=client_car, make_id=make_id if make_id else None)
        if form.is_valid():
            form.save()
            messages.success(request, 'Автомобиль клиента успешно обновлён!')
            return redirect('client_detail', pk=pk)
        else:
            print("Форма не валидна:", form.errors)  # Отладка
    else:
        form = ClientCarForm(instance=client_car, initial={'client': client})  # Задаем начальное значение client
    return render(request, 'clients/client_car_update.html', {
        'form': form,
        'client': client,
        'cancel_url': reverse('client_detail', kwargs={'pk': pk})
    })


# Удаление автомобиля клиента

def client_car_delete(request, pk, car_pk):
    client_car = get_object_or_404(ClientCar, pk=car_pk)
    client_pk = client_car.client.pk
    if request.method == 'POST':
        client_car.delete()
        messages.success(request, 'Автомобиль клиента успешно удалён!')
        return redirect('client_detail', pk=client_pk)
    return render(request, 'confirm_delete.html', {
        'object': client_car,
        'deleted_object_type': 'ClientCar',
        'cancel_url': reverse('client_detail', kwargs={'pk': client_pk})
    })

@login_required
def car_management(request):
    # Регистрируем функцию один раз для соединения
    register_custom_functions()

    car_makes = CarMake.objects.all()

    # Сортировка
    sort = request.GET.get('sort', 'id')
    direction = request.GET.get('direction', 'asc')
    if sort in ['id', 'name']:
        if direction == 'desc':
            car_makes = car_makes.order_by(f'-{sort}')
        else:
            car_makes = car_makes.order_by(sort)

    # Поиск
    search = request.GET.get('search', '')
    column = request.GET.get('column', 'name')

    if search:
        if column == 'id':
            car_makes = car_makes.filter(id__contains=search)
        elif column == 'name':
            search_lower = search.lower()
            car_makes = car_makes.extra(
                select={'name_lower': 'custom_lower(name)'},
                where=['custom_lower(name) LIKE %s'],
                params=[f'%{search_lower}%']
            )

    # Пагинация
    per_page = request.GET.get('per_page', 5)
    try:
        per_page = int(per_page)
        if per_page not in [5, 10, 20]:
            per_page = 5
    except ValueError:
        per_page = 5

    paginator = Paginator(car_makes, per_page)
    page_number = request.GET.get('page', 1)
    try:
        page_obj = paginator.get_page(page_number)
    except EmptyPage:
        page_obj = paginator.page(1)

    return render(request, 'cars/car_management.html', {
        'page_obj': page_obj,
        'per_page': per_page,
    })


# Создание марки автомобиля
def car_make_create(request):
    if request.method == 'POST':
        form = CarMakeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Марка автомобиля успешно создана!')
            return redirect('car_management')
    else:
        form = CarMakeForm()
    return render(request, 'cars/car_make_create.html', {'form': form, 'cancel_url': reverse('car_management')})


@login_required
def car_make_detail(request, pk):
    register_custom_functions()

    car_make = get_object_or_404(CarMake, pk=pk)
    models = CarModel.objects.filter(make=car_make)

    # Сортировка
    sort = request.GET.get('sort', 'id')
    direction = request.GET.get('direction', 'asc')
    if sort in ['id', 'name']:
        if direction == 'desc':
            models = models.order_by(f'-{sort}')
        else:
            models = models.order_by(sort)

    # Поиск
    search = request.GET.get('search', '')
    column = request.GET.get('column', 'name')

    if search:
        if column == 'id':
            models = models.filter(id__contains=search)
        elif column == 'name':
            search_lower = search.lower()
            models = models.extra(
                select={'name_lower': 'custom_lower(name)'},
                where=['custom_lower(name) LIKE %s'],
                params=[f'%{search_lower}%']
            )

    # Пагинация
    per_page = request.GET.get('per_page', 5)
    try:
        per_page = int(per_page)
        if per_page not in [5, 10, 20]:
            per_page = 5
    except ValueError:
        per_page = 5

    paginator = Paginator(models, per_page)
    page_number = request.GET.get('page', 1)
    try:
        page_obj = paginator.get_page(page_number)
    except EmptyPage:
        page_obj = paginator.page(1)

    return render(request, 'cars/car_make_detail.html', {
        'car_make': car_make,
        'page_obj': page_obj,
        'per_page': per_page,
        'cancel_url': reverse('car_make_detail', kwargs={'pk': pk})
    })

# Редактирование марки автомобиля
def car_make_update(request, pk):
    car_make = get_object_or_404(CarMake, pk=pk)
    if request.method == 'POST':
        form = CarMakeForm(request.POST, instance=car_make)
        if form.is_valid():
            form.save()
            messages.success(request, 'Марка автомобиля успешно обновлена!')
            return redirect('car_management')
    else:
        form = CarMakeForm(instance=car_make)
    return render(request, 'cars/car_make_update.html',
                  {'form': form, 'cancel_url': reverse('car_make_detail', kwargs={'pk': pk})})


# Удаление марки автомобиля


def car_make_delete(request, pk):
    car_make = get_object_or_404(CarMake, pk=pk)
    if request.method == 'POST':
        car_make.delete()
        messages.success(request, 'Марка автомобиля успешно удалена!')
        return redirect('car_management')
    return render(request, 'confirm_delete.html', {
        'object': car_make,
        'deleted_object_type': 'CarMake',
        'cancel_url': reverse('car_management')
    })


# Создание модели автомобиля для конкретной марки (без выпадающего списка марок)
def car_model_create(request, make_pk):
    car_make = get_object_or_404(CarMake, pk=make_pk)
    if request.method == 'POST':
        form = CarModelForm(request.POST)
        if form.is_valid():
            car_model = form.save(commit=False)
            car_model.make = car_make
            car_model.save()
            messages.success(request, f'Модель автомобиля успешно создана для марки {car_make.name}!')
            return redirect('car_make_detail', pk=make_pk)
    else:
        form = CarModelForm(initial={'make': car_make})
    return render(request, 'cars/car_model_create.html', {
        'form': form,
        'car_make': car_make,
        'cancel_url': reverse('car_make_detail', kwargs={'pk': make_pk})
    })


# Просмотр деталей модели автомобиля
def car_model_detail(request, pk):
    car_model = get_object_or_404(CarModel, pk=pk)
    return render(request, 'cars/car_model_detail.html',
                  {'car_model': car_model, 'cancel_url': reverse('car_make_detail', kwargs={'pk': car_model.make.pk})})


# Редактирование модели автомобиля
def car_model_update(request, pk):
    car_model = get_object_or_404(CarModel, pk=pk)
    if request.method == 'POST':
        form = CarModelForm(request.POST, instance=car_model)
        if form.is_valid():
            form.save()
            messages.success(request, 'Модель автомобиля успешно обновлена!')
            return redirect('car_make_detail', pk=car_model.make.pk)
    else:
        form = CarModelForm(instance=car_model)
    return render(request, 'cars/car_model_update.html', {
        'form': form,
        'car_make': car_model.make,
        'cancel_url': reverse('car_make_detail', kwargs={'pk': car_model.make.pk})
    })


# Удаление модели автомобиля


def car_model_delete(request, pk):
    car_model = get_object_or_404(CarModel, pk=pk)
    make_pk = car_model.make.pk
    if request.method == 'POST':
        car_model.delete()
        messages.success(request, 'Модель автомобиля успешно удалена!')
        return redirect('car_make_detail', pk=make_pk)
    return render(request, 'confirm_delete.html', {
        'object': car_model,
        'deleted_object_type': 'CarModel',
        'cancel_url': reverse('car_make_detail', kwargs={'pk': make_pk})
    })



@login_required
def service_management(request):
    # Регистрируем кастомную функцию для регистронезависимого поиска
    register_custom_functions()

    service_types = ServiceType.objects.all()

    # Сортировка
    sort = request.GET.get('sort', 'id')
    direction = request.GET.get('direction', 'asc')
    if sort in ['id', 'name']:
        if direction == 'desc':
            service_types = service_types.order_by(f'-{sort}')
        else:
            service_types = service_types.order_by(sort)

    # Поиск
    search = request.GET.get('search', '')
    column = request.GET.get('column', 'name')
    if search:
        if column == 'id':
            service_types = service_types.filter(id__contains=search)
        elif column == 'name':
            search_lower = search.lower()
            service_types = service_types.extra(
                select={'name_lower': 'custom_lower(name)'},
                where=['custom_lower(name) LIKE %s'],
                params=[f'%{search_lower}%']
            )

    # Пагинация
    per_page = request.GET.get('per_page', 5)
    try:
        per_page = int(per_page)
        if per_page not in [5, 10, 20]:
            per_page = 5
    except ValueError:
        per_page = 5

    paginator = Paginator(service_types, per_page)
    page_number = request.GET.get('page', 1)
    try:
        page_obj = paginator.get_page(page_number)
    except EmptyPage:
        page_obj = paginator.page(1)

    return render(request, 'services/service_management.html', {
        'page_obj': page_obj,
        'per_page': per_page
    })


# Создание типа услуги
def service_type_create(request):
    if request.method == 'POST':
        form = ServiceTypeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Тип услуги успешно создан!')
            return redirect('service_management')
    else:
        form = ServiceTypeForm()
    return render(request, 'services/service_type_create.html',
                  {'form': form, 'cancel_url': reverse('service_management')})


# Просмотр деталей типа услуги (с таблицей услуг и пагинацией)
@login_required
def service_type_detail(request, pk):
    # Регистрируем функцию один раз для соединения
    register_custom_functions()

    service_type = get_object_or_404(ServiceType, pk=pk)
    services = Service.objects.filter(service_type=service_type)

    # Сортировка
    sort = request.GET.get('sort', 'id')
    direction = request.GET.get('direction', 'asc')
    if sort in ['id', 'name']:
        if direction == 'desc':
            services = services.order_by(f'-{sort}')
        else:
            services = services.order_by(sort)

    # Поиск
    search = request.GET.get('search', '')
    column = request.GET.get('column', 'name')

    if search:
        if column == 'id':
            services = services.filter(id__contains=search)
        elif column == 'name':
            search_lower = search.lower()
            services = services.extra(
                select={'name_lower': 'custom_lower(name)'},
                where=['custom_lower(name) LIKE %s'],
                params=[f'%{search_lower}%']
            )

    # Пагинация
    per_page = request.GET.get('per_page', 5)
    try:
        per_page = int(per_page)
        if per_page not in [5, 10, 20]:
            per_page = 5
    except ValueError:
        per_page = 5

    paginator = Paginator(services, per_page)
    page_number = request.GET.get('page', 1)
    try:
        page_obj = paginator.get_page(page_number)
    except EmptyPage:
        page_obj = paginator.page(1)

    return render(request, 'services/service_type_detail.html', {
        'service_type': service_type,
        'page_obj': page_obj,
        'per_page': per_page,
        'cancel_url': reverse('service_management')
    })
# Редактирование типа услуги
def service_type_update(request, pk):
    service_type = get_object_or_404(ServiceType, pk=pk)
    if request.method == 'POST':
        form = ServiceTypeForm(request.POST, instance=service_type)
        if form.is_valid():
            form.save()
            messages.success(request, 'Тип услуги успешно обновлён!')
            return redirect('service_management')
    else:
        form = ServiceTypeForm(instance=service_type)
    return render(request, 'services/service_type_update.html',
                  {'form': form, 'cancel_url': reverse('service_management')})


# Удаление типа услуги
from .models import ServiceType


def service_type_delete(request, pk):
    from warehouse.models import WorkOrderService
    service_type = get_object_or_404(ServiceType, pk=pk)

    active_statuses = ('Первичный осмотр', 'Диагностика', 'В работе', 'Готов')
    active_wos = WorkOrderService.objects.filter(
        service__service_type=service_type,
        work_order__status__in=active_statuses,
    ).select_related('work_order')
    active_orders = list({w.work_order for w in active_wos})

    if request.method == 'POST':
        if active_orders:
            order_ids = ', '.join(f'№{o.id}' for o in active_orders)
            messages.error(
                request,
                f'Нельзя удалить тип услуги «{service_type.name}»: '
                f'его услуги используются в активных заказах {order_ids}.'
            )
            return redirect('service_management')
        service_type.delete()
        messages.success(request, 'Тип услуги успешно удалён!')
        return redirect('service_management')
    return render(request, 'confirm_delete.html', {
        'object': service_type,
        'deleted_object_type': 'ServiceType',
        'cancel_url': reverse('service_management'),
        'extra_warning': (
            f'Будут удалены все услуги этого типа ({service_type.service_set.count()} шт.). '
            'Данные завершённых заказов сохранятся (названия зафиксированы в снапшоте).'
        ) if not active_orders else None,
    })


# Создание конкретной услуги для типа услуги
def service_create(request, pk):  # Изменил type_pk на pk
    service_type = get_object_or_404(ServiceType, pk=pk)
    if request.method == 'POST':
        form = ServiceForm(request.POST)
        if form.is_valid():
            service = form.save(commit=False)
            service.service_type = service_type
            service.save()
            messages.success(request, f'Услуга "{service.name}" успешно создана для типа {service_type.name}!')
            return redirect('service_type_detail', pk=pk)
    else:
        form = ServiceForm(initial={'service_type': service_type})
    return render(request, 'services/service_create.html', {
        'form': form,
        'service_type': service_type,
        'cancel_url': reverse('service_type_detail', kwargs={'pk': pk})
    })


# Просмотр деталей услуги
def service_detail(request, pk, service_pk):  # Изменил type_pk на pk для соответствия маршруту
    service = get_object_or_404(Service, pk=service_pk)  # Используем service_pk для получения услуги
    return render(request, 'services/service_detail.html',
                  {'service': service, 'cancel_url': reverse('service_type_detail', kwargs={'pk': pk})})


# Удаление услуги

def service_delete(request, pk, service_pk):
    service = get_object_or_404(Service, pk=service_pk)
    if request.method == 'POST':
        service.delete()
        messages.success(request, 'Услуга успешно удалена!')
        return redirect('service_type_detail', pk=pk)
    return render(request, 'confirm_delete.html', {
        'object': service,
        'deleted_object_type': 'Service',
        'cancel_url': reverse('service_type_detail', kwargs={'pk': pk})
    })


# Редактирование услуги
def service_update(request, pk, service_pk):  # Изменил на pk (ID типа услуги) и service_pk (ID услуги)
    service = get_object_or_404(Service, pk=service_pk)  # Используем service_pk для получения услуги
    if request.method == 'POST':
        form = ServiceForm(request.POST, instance=service)
        if form.is_valid():
            form.save()
            messages.success(request, 'Услуга успешно обновлена!')
            return redirect('service_type_detail', pk=pk)  # Используем pk для возврата к типу услуги
    else:
        form = ServiceForm(instance=service)
    return render(request, 'services/service_update.html', {
        'form': form,
        'service_type': service.service_type,  # Передаём тип услуги для контекста
        'cancel_url': reverse('service_type_detail', kwargs={'pk': pk})
    })






from .models import Order

@login_required
def orders_list(request):
    # Регистрируем функцию для регистронезависимого поиска
    register_custom_functions()

    orders = Order.objects.all()

    # Сортировка
    sort = request.GET.get('sort', 'id')
    direction = request.GET.get('direction', 'desc')
    valid_sort_fields = ['id', 'client_fio_static', 'car_details_static', 'order_date', 'status']

    if sort == 'client_fio_static':
        # Регистронезависимая сортировка по ФИО
        orders = orders.extra(select={'client_lower': 'custom_lower(client_fio_static)'}).order_by(
            f"{'-' if direction == 'desc' else ''}client_lower"
        )
    elif sort == 'car_details_static':
        # Регистронезависимая сортировка по авто
        orders = orders.extra(select={'car_lower': 'custom_lower(car_details_static)'}).order_by(
            f"{'-' if direction == 'desc' else ''}car_lower"
        )
    elif sort in valid_sort_fields:
        # Обычная сортировка по остальным полям
        orders = orders.order_by(f'-{sort}' if direction == 'desc' else sort)

    # Поиск
    search = request.GET.get('search', '')
    column = request.GET.get('column', 'client')

    if search:
        search_lower = search.lower()
        if column == 'id':
            orders = orders.filter(id__contains=search)
        elif column == 'client':
            orders = orders.extra(
                select={'client_lower': 'custom_lower(client_fio_static)'},
                where=['custom_lower(client_fio_static) LIKE %s'],
                params=[f'%{search_lower}%']
            )
        elif column == 'car':
            orders = orders.extra(
                select={'car_lower': 'custom_lower(car_details_static)'},
                where=['custom_lower(car_details_static) LIKE %s'],
                params=[f'%{search_lower}%']
            )

    # Пагинация
    per_page = request.GET.get('per_page', 5)
    try:
        per_page = int(per_page)
        if per_page not in [5, 10, 20]:
            per_page = 5
    except ValueError:
        per_page = 5

    paginator = Paginator(orders, per_page)
    page_number = request.GET.get('page', 1)
    try:
        page_obj = paginator.get_page(page_number)
    except EmptyPage:
        page_obj = paginator.page(1)

    return render(request, 'orders/orders_list.html', {
        'page_obj': page_obj,
        'per_page': per_page,
    })

# ── Order helpers ─────────────────────────────────────────────

def _entries_cheapest_first(part):
    """Return StockEntry objects sorted cheapest-first, matching _reserve_cheapest_first order.

    _reserve_cheapest_first places reservations on the cheapest location first.
    When releasing/issuing we must visit entries in the same order so we always
    touch the entry where the reservation actually sits, not a random sibling.
    """
    from warehouse.models import StockEntry
    from warehouse.views import _in_stock_batches
    from decimal import Decimal as _D
    loc_min_price = {}
    for b in _in_stock_batches(part, available_only=False):
        lid = b['location'].id
        if lid not in loc_min_price or b['price_per_unit'] < loc_min_price[lid]:
            loc_min_price[lid] = b['price_per_unit']
    entries = list(StockEntry.objects.filter(part=part).select_related('location'))
    entries.sort(key=lambda e: (loc_min_price.get(e.location_id, _D('999999')), e.id))
    return entries


def _release_wop_reservation(wop):
    """Release stock reservation for a WOP using tracked entry list (or cheapest-first fallback)."""
    from warehouse.models import StockEntry
    tracked = wop.reserved_entries or []
    if tracked:
        for item in tracked:
            entry = StockEntry.objects.filter(id=item['entry_id']).first()
            if entry:
                entry.reserved_qty -= min(entry.reserved_qty, item['qty'])
                entry.save()
    elif wop.part:
        remaining = wop.quantity
        for entry in _entries_cheapest_first(wop.part):
            if remaining <= 0:
                break
            release = min(entry.reserved_qty, remaining)
            entry.reserved_qty -= release
            entry.save()
            remaining -= release


def _unreserve_order_parts(order):
    """Release stock reservations for all reserved parts in this order."""
    from warehouse.models import WorkOrderPart
    for wop in WorkOrderPart.objects.filter(work_order=order, status='reserved'):
        _release_wop_reservation(wop)
        wop.status = 'cancelled'
        wop.save()


def _issue_order_parts(order):
    """Move reserved parts out of stock when an order is completed."""
    from warehouse.models import WorkOrderPart, StockEntry
    for wop in WorkOrderPart.objects.filter(work_order=order, status='reserved'):
        tracked = wop.reserved_entries or []
        if tracked:
            for item in tracked:
                entry = StockEntry.objects.filter(id=item['entry_id']).first()
                if entry:
                    take = min(entry.reserved_qty, item['qty'])
                    entry.reserved_qty -= take
                    entry.total_qty -= take
                    entry.save()
        else:
            remaining = wop.quantity
            for entry in _entries_cheapest_first(wop.part):
                if remaining <= 0:
                    break
                take = min(entry.reserved_qty, remaining)
                entry.reserved_qty -= take
                entry.total_qty -= take
                entry.save()
                remaining -= take
        wop.status = 'cancelled'
        wop.save()


def _freeze_service_snapshots(order):
    """Ensure all WorkOrderService records for this order have snapshots saved."""
    from warehouse.models import WorkOrderService, WorkshopSettings
    settings = WorkshopSettings.objects.first()
    current_rate = settings.hourly_rate if settings else _Decimal('0')
    for wos in WorkOrderService.objects.filter(work_order=order).select_related('service'):
        changed = False
        if not wos.service_name_snapshot:
            wos.service_name_snapshot = wos.service.name if wos.service else '—'
            changed = True
        if wos.hourly_rate_snapshot is None:
            wos.hourly_rate_snapshot = current_rate
            changed = True
        if changed:
            wos.save(update_fields=['service_name_snapshot', 'hourly_rate_snapshot'])


def _recalculate_order_cost(order):
    from warehouse.models import WorkOrderService, WorkOrderPart
    wos_total = sum(
        (w.final_price or _Decimal('0')) for w in WorkOrderService.objects.filter(work_order=order)
    )
    wop_total = sum(
        (w.sale_price or _Decimal('0'))
        for w in WorkOrderPart.objects.filter(work_order=order)
    )
    order.cost = wos_total + wop_total
    order.save(update_fields=['cost'])


def _unreserve_single_wop(wop):
    """Release stock reservation for one WorkOrderPart (reserved status only)."""
    if wop.status != 'reserved':
        return
    _release_wop_reservation(wop)


@login_required
def order_commit(request, pk):
    """Batch-save all staged changes from the order detail page in one atomic operation."""
    import json as _json_mod
    from django.db import transaction
    from warehouse.models import (
        WorkOrderService, WorkOrderPart, WorkOrderServiceEmployee,
        StockEntry, WorkshopSettings, Employee,
    )
    from warehouse.views import _min_stock_price, _reserve_cheapest_first
    from mainapp.models import Service
    from mainapp.signals import _locals as _sig_locals, _create_log

    if request.method != 'POST':
        return _json_response({'error': 'Method not allowed'}, 405)

    order = get_object_or_404(Order, pk=pk)
    if order.status in ('Завершён', 'Отменён'):
        return _json_response({'error': 'Редактирование закрытого заказа запрещено'}, 400)

    old_cost = order.cost  # capture before transaction for cost-change logging

    try:
        data = _json_mod.loads(request.body)
    except (ValueError, _json_mod.JSONDecodeError):
        return _json_response({'error': 'Неверный формат запроса'}, 400)

    # Optimistic locking
    client_version = data.get('client_version', '')
    if client_version and client_version != order.updated_at.isoformat():
        return _json_response({
            'conflict': True,
            'message': 'Заказ был изменён другим сотрудником. Страница будет обновлена.',
        }, 409)

    errors = []
    warnings = []
    log_field_changes = []
    log_events = []

    _sig_locals.suppress_signals = True
    try:
        with transaction.atomic():
            # 1. Remove services (unreserve their parts, then delete cascade)
            for wos_id in data.get('remove_services', []):
                wos = WorkOrderService.objects.filter(pk=wos_id, work_order=order).first()
                if wos:
                    svc_name = wos.service_name_snapshot or (wos.service.name if wos.service else '—')
                    for wop in wos.parts.filter(status='reserved'):
                        _unreserve_single_wop(wop)
                    wos.parts.all().delete()
                    wos.delete()
                    log_events.append({'event': 'service_removed', 'service': svc_name})

            # 2. Remove standalone parts
            for wop_id in data.get('remove_parts', []):
                wop = WorkOrderPart.objects.filter(pk=wop_id, work_order=order).first()
                if wop:
                    log_events.append({
                        'event': 'part_removed',
                        'article': wop.part.article if wop.part else '—',
                        'part_name': wop.part.name if wop.part else '—',
                        'quantity': str(wop.quantity),
                    })
                    _unreserve_single_wop(wop)
                    wop.delete()

            # 3. Unassign employees
            for item in data.get('unassign_employees', []):
                for a in WorkOrderServiceEmployee.objects.filter(
                    work_order_service_id=item.get('wosId'),
                    employee_id=item.get('empId'),
                ).select_related('employee', 'work_order_service__service'):
                    log_events.append({
                        'event': 'employee_removed',
                        'employee': a.employee.name if a.employee else '—',
                        'service': (a.work_order_service.service_name_snapshot or
                                    (a.work_order_service.service.name if a.work_order_service.service else '—')),
                    })
                WorkOrderServiceEmployee.objects.filter(
                    work_order_service_id=item.get('wosId'),
                    employee_id=item.get('empId'),
                ).delete()

            # 4. Assign employees (guard: worker may have been deleted)
            for item in data.get('assign_employees', []):
                emp = Employee.objects.filter(pk=item.get('empId')).first()
                if not emp:
                    warnings.append(f'Сотрудник ID={item.get("empId")} не найден — возможно, был удалён.')
                    continue
                wos = WorkOrderService.objects.filter(pk=item.get('wosId'), work_order=order).first()
                if not wos:
                    warnings.append(f'Услуга ID={item.get("wosId")} не найдена.')
                    continue
                asgn, created = WorkOrderServiceEmployee.objects.get_or_create(
                    work_order_service=wos, employee=emp,
                    defaults={
                        'salary_coefficient_snapshot': emp.salary_coefficient,
                        'employee_name_snapshot': emp.name,
                    },
                )
                if not created:
                    update_fields = []
                    if asgn.salary_coefficient_snapshot is None:
                        asgn.salary_coefficient_snapshot = emp.salary_coefficient
                        update_fields.append('salary_coefficient_snapshot')
                    if not asgn.employee_name_snapshot:
                        asgn.employee_name_snapshot = emp.name
                        update_fields.append('employee_name_snapshot')
                    if update_fields:
                        asgn.save(update_fields=update_fields)
                if created:
                    log_events.append({
                        'event': 'employee_assigned',
                        'employee': emp.name,
                        'service': wos.service_name_snapshot or (wos.service.name if wos.service else '—'),
                    })

            # 5. Add services
            temp_wos_map = {}  # tempId (str) -> WorkOrderService instance
            settings = WorkshopSettings.objects.first()
            rate = settings.hourly_rate if settings else _Decimal('0')
            for svc_data in data.get('add_services', []):
                svc = Service.objects.filter(pk=svc_data.get('id')).first()
                if not svc:
                    errors.append(f'Услуга ID={svc_data.get("id")} не найдена.')
                    continue
                hours = _Decimal(str(svc_data.get('hours', svc.base_hours)))
                factor = _Decimal(str(svc_data.get('factor', '1.0')))
                final_price = (hours * rate * factor).quantize(_Decimal('0.01'))
                wos = WorkOrderService.objects.create(
                    work_order=order,
                    service=svc,
                    service_name_snapshot=svc.name,
                    hourly_rate_snapshot=rate,
                    hours_applied=hours,
                    complexity_factor=factor,
                )
                temp_id = svc_data.get('tempId')
                if temp_id is not None:
                    temp_wos_map[str(temp_id)] = wos
                for emp_id in svc_data.get('employees', []):
                    emp = Employee.objects.filter(pk=emp_id).first()
                    if emp:
                        WorkOrderServiceEmployee.objects.get_or_create(
                            work_order_service=wos,
                            employee=emp,
                            defaults={
                                'salary_coefficient_snapshot': emp.salary_coefficient,
                                'employee_name_snapshot': emp.name,
                            },
                        )
                log_events.append({
                    'event': 'service_added',
                    'service': svc.name,
                    'hours': str(hours),
                    'complexity': str(factor),
                    'price': str(final_price),
                })

            # 6. Add parts (guard: part may be deleted or stock exhausted)
            for pd in data.get('add_parts', []):
                from warehouse.models import Part
                part = Part.objects.filter(pk=pd.get('partId')).first()
                if not part:
                    errors.append(f'Деталь ID={pd.get("partId")} не найдена — возможно, была удалена.')
                    continue
                qty = int(pd.get('qty', 1))
                markup = _Decimal(str(pd.get('markup', '30')))
                total_avail = sum(e.available_qty for e in StockEntry.objects.filter(part=part))
                if total_avail == 0:
                    errors.append(f'{part.article} «{part.name}»: нет в наличии на складе.')
                    continue
                if total_avail < qty:
                    warnings.append(
                        f'{part.article}: запрошено {qty} шт., доступно {total_avail} шт. '
                        f'Зарезервировано {total_avail} шт.'
                    )
                    qty = total_avail
                cost_total = _min_stock_price(part, qty)
                sale_price = (cost_total * (1 + markup / _Decimal('100'))).quantize(_Decimal('0.01')) if cost_total else None
                markup = markup.quantize(_Decimal('0.01'))
                expected_price = pd.get('expectedPrice')
                if expected_price is not None and sale_price is not None:
                    exp = _Decimal(str(expected_price)).quantize(_Decimal('0.01'))
                    if abs(sale_price - exp) > _Decimal('0.01'):
                        warnings.append(
                            f'Цена {part.article} «{part.name}» изменилась при сохранении: '
                            f'ожидалось {exp} руб., фактически {sale_price} руб. '
                            f'(складские остатки изменились пока вы работали).'
                        )
                temp_wos_id = pd.get('tempWosId')
                wos_id = pd.get('wosId')
                if temp_wos_id is not None:
                    wos = temp_wos_map.get(str(temp_wos_id))
                else:
                    wos = WorkOrderService.objects.filter(pk=wos_id, work_order=order).first() if wos_id else None
                shortage, res_entries = _reserve_cheapest_first(part, qty)
                WorkOrderPart.objects.create(
                    work_order=order, part=part, quantity=qty,
                    markup=markup, sale_price=sale_price,
                    work_order_service=wos, status='reserved',
                    reserved_entries=res_entries,
                    part_article_snapshot=part.article or '',
                    part_name_snapshot=part.name or '',
                )
                log_events.append({
                    'event': 'part_added',
                    'article': part.article,
                    'part_name': part.name,
                    'quantity': str(qty),
                    'sale_price': str(sale_price) if sale_price else '—',
                })

            # 7. Update services (hours / factor)
            for item in data.get('update_services', []):
                wos = WorkOrderService.objects.filter(pk=item.get('wosId'), work_order=order).first()
                if not wos:
                    errors.append(f'Услуга ID={item.get("wosId")} не найдена.')
                    continue
                new_hours  = _Decimal(str(item.get('hours',  wos.hours_applied)))
                new_factor = _Decimal(str(item.get('factor', wos.complexity_factor)))
                wos.hours_applied      = new_hours
                wos.complexity_factor  = new_factor
                wos.save()
                log_events.append({
                    'event': 'service_updated',
                    'service': wos.service_name_snapshot or (wos.service.name if wos.service else '—'),
                    'hours': str(new_hours),
                    'complexity': str(new_factor),
                })

            # 8. Update parts (qty / markup)
            for item in data.get('update_parts', []):
                wop = WorkOrderPart.objects.filter(pk=item.get('wopId'), work_order=order).first()
                if not wop:
                    errors.append(f'Деталь ID={item.get("wopId")} не найдена.')
                    continue
                new_qty    = int(item.get('qty',    wop.quantity))
                new_markup = _Decimal(str(item.get('markup', wop.markup))).quantize(_Decimal('0.01'))
                delta = new_qty - wop.quantity
                extra_cost = None
                if delta != 0 and wop.status == 'reserved':
                    if delta > 0:
                        # Price BEFORE reserving so _min_stock_price sees available batches
                        extra_cost = _min_stock_price(wop.part, delta)
                        shortage, new_res = _reserve_cheapest_first(wop.part, delta)
                        if shortage > 0:
                            warnings.append(f'{wop.part.article}: не удалось зарезервировать {shortage} шт.')
                            actually_reserved = delta - shortage
                            new_qty -= shortage
                            if extra_cost and actually_reserved > 0:
                                extra_cost = extra_cost * _Decimal(actually_reserved) / _Decimal(delta)
                            else:
                                extra_cost = _Decimal('0')
                        # Merge new reservation tracking into wop
                        tracked = list(wop.reserved_entries or [])
                        for nr in new_res:
                            merged = False
                            for ex in tracked:
                                if ex['entry_id'] == nr['entry_id']:
                                    ex['qty'] += nr['qty']
                                    merged = True
                                    break
                            if not merged:
                                tracked.append(nr)
                        wop.reserved_entries = tracked
                    else:
                        # Release excess — remove from most-expensive end (last in list)
                        to_release = -delta
                        tracked = list(wop.reserved_entries or [])
                        new_tracked = []
                        for item in reversed(tracked):
                            if to_release <= 0:
                                new_tracked.insert(0, item)
                                continue
                            entry = StockEntry.objects.filter(id=item['entry_id']).first()
                            release = min(item['qty'], to_release)
                            if entry and release > 0:
                                entry.reserved_qty -= release
                                entry.save()
                            to_release -= release
                            remaining_in_item = item['qty'] - release
                            if remaining_in_item > 0:
                                new_tracked.insert(0, {'entry_id': item['entry_id'], 'qty': remaining_in_item})
                        wop.reserved_entries = new_tracked
                if (new_markup != wop.markup or new_qty != wop.quantity) and wop.sale_price and wop.quantity:
                    old_factor = 1 + wop.markup / _Decimal('100')
                    old_cost = wop.sale_price / old_factor  # total purchase cost without markup
                    old_qty = wop.quantity
                    if new_qty > old_qty:
                        total_cost = old_cost + (extra_cost or _Decimal('0'))
                    elif new_qty < old_qty:
                        total_cost = old_cost * _Decimal(new_qty) / _Decimal(old_qty)
                    else:
                        total_cost = old_cost
                    wop.sale_price = (
                        total_cost * (1 + new_markup / _Decimal('100'))
                    ).quantize(_Decimal('0.01'))
                wop.quantity = new_qty
                wop.markup   = new_markup
                wop.save(update_fields=['quantity', 'markup', 'sale_price', 'reserved_entries'])
                log_events.append({
                    'event': 'part_updated',
                    'article': wop.part.article if wop.part else '—',
                    'part_name': wop.part.name if wop.part else '—',
                    'quantity': str(new_qty),
                    'markup': str(new_markup),
                })

            # 9. Update order status / comment / payment / mileage
            order_upd = data.get('order_update')
            if order_upd:
                # Block all field changes on completed orders
                if order.status == 'Завершён':
                    errors.append('Заказ завершён — изменение статуса и параметров запрещено.')
                    order_upd = None

            if order_upd:
                new_status = order_upd.get('status', order.status)
                new_comment = order_upd.get('comment', order.comment or '')
                if order.status != new_status:
                    log_field_changes.append({'field': 'Статус', 'from': order.status, 'to': new_status})
                if (order.comment or '') != new_comment:
                    log_field_changes.append({
                        'field': 'Комментарий',
                        'from': order.comment or '—',
                        'to': new_comment or '—',
                    })

                # Payment method — only allowed when order is/becomes 'Готов'
                effective_status = new_status
                if effective_status != 'Готов' and order.status != 'Готов':
                    order_upd['payment_method'] = order.payment_method  # ignore submitted value

                new_payment = order_upd.get('payment_method', order.payment_method)
                if new_payment != order.payment_method:
                    payment_labels = dict(Order.PAYMENT_METHOD_CHOICES)
                    log_field_changes.append({
                        'field': 'Форма оплаты',
                        'from': payment_labels.get(order.payment_method, order.payment_method or '—'),
                        'to': payment_labels.get(new_payment, new_payment or '—'),
                    })
                order.payment_method = new_payment or None

                # Mileage
                raw_mileage = order_upd.get('mileage')
                if raw_mileage is not None and raw_mileage != '':
                    try:
                        new_mileage = int(raw_mileage)
                    except (ValueError, TypeError):
                        new_mileage = order.mileage
                else:
                    new_mileage = order.mileage

                if new_mileage is not None and new_mileage != order.mileage:
                    mileage_reason = (order_upd.get('mileage_change_reason') or '').strip()
                    log_field_changes.append({
                        'field': 'Пробег',
                        'from': str(order.mileage) if order.mileage is not None else '—',
                        'reason': mileage_reason or '—',
                        'to': str(new_mileage),
                    })
                    if order.mileage is not None:
                        order.mileage_prev = order.mileage
                        order.mileage_change_reason = mileage_reason or None
                order.mileage = new_mileage

                if new_status == 'Завершён' and order.status != 'Готов':
                    errors.append(
                        'Нельзя завершить заказ: текущий статус должен быть «Готов». '
                        'Сначала переведите заказ в статус «Готов» и укажите форму оплаты.'
                    )
                    new_status = order.status  # revert
                elif new_status == 'Завершён' and not (order.payment_method or new_payment):
                    errors.append(
                        'Нельзя завершить заказ: не указана форма оплаты. '
                        'Укажите форму оплаты в статусе «Готов» перед завершением.'
                    )
                    new_status = order.status  # revert

                if new_status == 'Отменён':
                    _unreserve_order_parts(order)
                    _freeze_service_snapshots(order)
                elif new_status == 'Завершён':
                    _freeze_service_snapshots(order)
                    _issue_order_parts(order)
                    if order.client_car:
                        order.client_fio_static = order.client_fio_static or order.client_car.client.fio
                        order.car_details_static = order.car_details_static or (
                            f"{order.client_car.make.name} {order.client_car.model.name} "
                            f"({order.client_car.license_plate or ''}"
                            f"{', VIN: ' + order.client_car.vin if order.client_car.vin else ''})"
                        )
                        wos_sum = "; ".join(
                            f"{w.service_name_snapshot or (w.service.name if w.service else '—')} ({w.final_price} руб.)"
                            for w in WorkOrderService.objects.filter(work_order=order).select_related('service')
                        )
                        wop_sum = "; ".join(
                            f"{w.part_article_snapshot or (w.part.article if w.part else '—')} × {w.quantity}"
                            for w in WorkOrderPart.objects.filter(work_order=order).select_related('part')
                        )
                        order.services_static = wos_sum or "Нет услуг"
                        order.custom_services_static = wop_sum or "Нет деталей"
                order.status = new_status
                order.comment = new_comment
                order.save()

            _recalculate_order_cost(order)
            # Always bump updated_at so concurrent editors detect the conflict.
            # _recalculate_order_cost uses update_fields=['cost'] which skips auto_now.
            from django.utils import timezone as _tz
            Order.objects.filter(pk=order.pk).update(updated_at=_tz.now())

            new_cost = order.cost
            if old_cost != new_cost:
                log_field_changes.append({
                    'field': 'Стоимость',
                    'from': str(old_cost) if old_cost is not None else '—',
                    'to': str(new_cost) if new_cost is not None else '—',
                })
    finally:
        _sig_locals.suppress_signals = False

    # One consolidated log entry for this commit (only what actually changed)
    if log_field_changes or log_events:
        _create_log('update', 'Order', order.pk, order,
                    {'field_changes': log_field_changes, 'events': log_events})

    return _json_response({'success': not bool(errors), 'errors': errors, 'warnings': warnings})


def _json_response(data, status=200):
    from django.http import JsonResponse
    return JsonResponse(data, status=status)


# ── Order CRUD ────────────────────────────────────────────────

@login_required
def order_create(request):
    from warehouse.models import WorkshopSettings
    if request.method == 'POST':
        client_form = ClientSelectionForm(request.POST)
        order_form = OrderForm(request.POST)
        if client_form.is_valid() and order_form.is_valid():
            order = order_form.save(commit=False)
            ws = WorkshopSettings.objects.first()
            if ws:
                order.org_snapshot = ws.as_snapshot()
            order.save()
            messages.success(request, 'Заказ создан.')
            return redirect('order_detail', pk=order.pk)
    else:
        client_form = ClientSelectionForm()
        order_form = OrderForm()
    return render(request, 'orders/order_create.html', {
        'client_form': client_form,
        'order_form': order_form,
        'cancel_url': reverse('orders_list'),
    })


@login_required
def order_detail(request, pk):
    from warehouse.models import WorkOrderService, WorkOrderPart, WorkshopSettings
    from mainapp.models import ServiceType, Service as _Service
    order = get_object_or_404(Order, pk=pk)

    wos_list = list(order.work_order_services.select_related('service').prefetch_related(
        'parts__part', 'assignments__employee'
    ).all())
    wop_list = list(order.work_order_parts.select_related(
        'part', 'work_order_service__service'
    ).all())

    services_total = sum((w.final_price or _Decimal('0')) for w in wos_list)
    parts_total = sum((w.sale_price or _Decimal('0')) for w in wop_list)
    total = services_total + parts_total

    if order.status not in ('Завершён', 'Отменён') and order.cost != total:
        order.cost = total
        order.save(update_fields=['cost'])

    settings = WorkshopSettings.objects.first()

    grouped_services = []
    for stype in ServiceType.objects.prefetch_related('service_set').all():
        svcs = list(stype.service_set.all())
        if svcs:
            grouped_services.append((stype.name, svcs))

    service_hours = {s.id: float(s.base_hours) for stype_name, svcs in grouped_services for s in svcs}

    unlinked_wop_list = [w for w in wop_list if not w.work_order_service]

    # Detect if live car/client data differs from snapshot (to show a note in the card)
    car_snap_live = None
    if order.car_snapshot and order.client_car:
        cs = order.car_snapshot
        cc = order.client_car
        live = {
            'client_fio':   cc.client.fio,
            'client_phone': cc.client.phone or '',
            'make':         cc.make.name,
            'model':        cc.model.name,
            'plate':        cc.license_plate or '',
            'vin':          cc.vin or '',
        }
        changed = [k for k in live if cs.get(k, '') != live[k]]
        if changed:
            car_snap_live = {
                'client': f'{live["client_fio"]}' + (f', {live["client_phone"]}' if live['client_phone'] else ''),
                'car':    f'{live["make"]} {live["model"]}' + (f' ({live["plate"]})' if live['plate'] else ''),
            }

    return render(request, 'orders/order_detail.html', {
        'order': order,
        'wos_list': wos_list,
        'wop_list': wop_list,
        'unlinked_wop_list': unlinked_wop_list,
        'services_total': services_total,
        'parts_total': parts_total,
        'total': total,
        'hourly_rate': settings.hourly_rate if settings else 0,
        'grouped_services': grouped_services,
        'service_hours_json': _json.dumps(service_hours),
        'editable': order.status not in ('Завершён', 'Отменён'),
        'order_version': order.updated_at.isoformat(),
        'order_status_choices': Order.STATUS_CHOICES,
        'payment_method_choices': Order.PAYMENT_METHOD_CHOICES,
        'car_snap_live': car_snap_live,
    })


@login_required
def order_update(request, pk):
    order = get_object_or_404(Order, pk=pk)

    if order.status in ('Завершён', 'Отменён'):
        messages.error(request, 'Редактирование завершённого или отменённого заказа запрещено.')
        return redirect('order_detail', pk=pk)

    if request.method == 'POST':
        # Optimistic locking: detect concurrent edits
        client_version = request.POST.get('client_version', '')
        if client_version and client_version != order.updated_at.isoformat():
            messages.warning(
                request,
                'Этот заказ был изменён другим сотрудником пока вы работали. '
                'Страница обновлена — проверьте актуальное состояние перед сохранением.'
            )
            return redirect('order_detail', pk=pk)

        form = OrderForm(request.POST, instance=order)
        if form.is_valid():
            new_status = form.cleaned_data['status']
            order_obj = form.save(commit=False)

            if new_status == 'Отменён':
                _unreserve_order_parts(order)
                _freeze_service_snapshots(order)
            elif new_status == 'Завершён':
                _freeze_service_snapshots(order)
                _issue_order_parts(order)
                if order.client_car:
                    order_obj.client_fio_static = order_obj.client_fio_static or order.client_car.client.fio
                    order_obj.car_details_static = order_obj.car_details_static or (
                        f"{order.client_car.make.name} {order.client_car.model.name} "
                        f"({order.client_car.license_plate or ''}"
                        f"{', VIN: ' + order.client_car.vin if order.client_car.vin else ''})"
                    )
                    from warehouse.models import WorkOrderService as _WOS, WorkOrderPart as _WOP
                    wos_sum = "; ".join(
                        f"{w.service_name_snapshot or (w.service.name if w.service else '—')} ({w.final_price} руб.)"
                        for w in _WOS.objects.filter(work_order=order).select_related('service')
                    )
                    wop_sum = "; ".join(
                        f"{w.part_article_snapshot or (w.part.article if w.part else '—')} × {w.quantity}"
                        for w in _WOP.objects.filter(work_order=order).select_related('part')
                    )
                    order_obj.services_static = wos_sum or "Нет услуг"
                    order_obj.custom_services_static = wop_sum or "Нет деталей"

            order_obj.save()
            _recalculate_order_cost(order_obj)
            messages.success(request, 'Заказ обновлён.')
            return redirect('order_detail', pk=pk)
    else:
        form = OrderForm(instance=order)

    return render(request, 'orders/order_update.html', {
        'order_form': form,
        'order': order,
        'cancel_url': reverse('order_detail', kwargs={'pk': pk}),
    })


@login_required
def order_delete(request, pk):
    order = get_object_or_404(Order, pk=pk)
    if order.status in ('Завершён', 'Отменён'):
        messages.error(
            request,
            f'Нельзя удалить заказ со статусом «{order.status}». '
            'Завершённые и отменённые заказы хранятся для истории.'
        )
        return redirect('order_detail', pk=pk)
    if request.method == 'POST':
        _unreserve_order_parts(order)
        order.delete()
        messages.success(request, 'Заказ удалён.')
        return redirect('orders_list')
    return render(request, 'confirm_delete.html', {
        'object': order,
        'deleted_object_type': 'Order',
        'cancel_url': reverse('order_detail', kwargs={'pk': pk}),
    })


from django.shortcuts import render

def help_page(request):
    return render(request, 'help/help.html')


# Кастомный декоратор для проверки прав администратора
def staff_required(view_func):
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated or not request.user.is_staff:
            messages.error(request, "У вас нет прав для выполнения этого действия.")
            return redirect('orders_list')  # Перенаправляем не-админов на список заказов
        return view_func(request, *args, **kwargs)
    return wrapper

@login_required
@staff_required
def accounting_export(request):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from decimal import Decimal
    import io
    from datetime import date, datetime
    from warehouse.models import WorkOrderService, WorkOrderPart, SupplyItem, WorkOrderServiceEmployee

    today = date.today()
    first_of_month = today.replace(day=1)

    params = request.POST if request.method == 'POST' else request.GET
    date_from_str = params.get('date_from', first_of_month.strftime('%Y-%m-%d'))
    date_to_str   = params.get('date_to',   today.strftime('%Y-%m-%d'))

    try:
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
        date_to   = datetime.strptime(date_to_str,   '%Y-%m-%d').date()
    except ValueError:
        date_from = first_of_month
        date_to   = today

    if request.method != 'POST':
        return render(request, 'admin/accounting_export.html', {
            'date_from': date_from_str,
            'date_to':   date_to_str,
        })

    # ── Стили ──────────────────────────────────────────────────────────────
    BLUE     = 'FF1E40AF'
    LBLUE    = 'FFF0F4FF'
    WHITE    = 'FFFFFFFF'
    TOTAL_BG = 'FFE2E8F0'
    WARN_BG  = 'FFFFF3CD'

    hdr_font  = Font(bold=True, color='FFFFFFFF', name='Calibri', size=10)
    data_font = Font(name='Calibri', size=10)
    bold_font = Font(bold=True, name='Calibri', size=10)
    thin_side = Side(style='thin', color='FFCBD5E1')
    thin_border = Border(left=thin_side, right=thin_side,
                         top=thin_side, bottom=thin_side)

    def hdr_fill():  return PatternFill('solid', fgColor=BLUE)
    def alt_fill(i): return PatternFill('solid', fgColor=LBLUE if i % 2 == 0 else WHITE)
    def tot_fill():  return PatternFill('solid', fgColor=TOTAL_BG)

    def style_header(ws, headers, col_widths):
        ws.append(headers)
        for col, w in enumerate(col_widths, 1):
            cell = ws.cell(1, col)
            cell.fill   = hdr_fill()
            cell.font   = hdr_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = 'A2'

    def style_row(ws, row_idx, n_cols, is_money_cols=None):
        fill = alt_fill(row_idx)
        for col in range(1, n_cols + 1):
            cell = ws.cell(row_idx + 1, col)
            cell.fill   = fill
            cell.font   = data_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')
            if is_money_cols and col in is_money_cols:
                cell.number_format = '#,##0.00'

    def add_total_row(ws, row_idx, n_cols, sum_cols, label_col=1, label='ИТОГО'):
        ws.cell(row_idx + 1, label_col).value = label
        for col in range(1, n_cols + 1):
            cell = ws.cell(row_idx + 1, col)
            cell.fill   = tot_fill()
            cell.font   = bold_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')
            if col in sum_cols:
                cell.number_format = '#,##0.00'

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # убираем Sheet по умолчанию

    period_label = f'{date_from.strftime("%d.%m.%Y")} – {date_to.strftime("%d.%m.%Y")}'

    # ══════════════════════════════════════════════════════════════════════
    # Лист 1: Реализация (завершённые заказы)
    # ══════════════════════════════════════════════════════════════════════
    ws1 = wb.create_sheet('Реализация')
    headers1 = [
        'Дата завершения', '№ заказа', 'Клиент', 'Телефон клиента',
        'Автомобиль', 'Гос. номер', 'Сумма услуг (руб.)',
        'Сумма запчастей (руб.)', 'ИТОГО (руб.)',
    ]
    widths1 = [16, 10, 28, 18, 24, 14, 20, 22, 16]
    style_header(ws1, headers1, widths1)

    orders = Order.objects.filter(
        status='Завершён',
        completion_date__gte=date_from,
        completion_date__lte=date_to,
    ).order_by('completion_date')

    total_svcs = Decimal('0'); total_parts_sum = Decimal('0'); total_all = Decimal('0')
    for i, o in enumerate(orders):
        svcs_sum  = sum(w.final_price or Decimal('0') for w in o.work_order_services.all())
        parts_sum = sum((w.sale_price or Decimal('0')) for w in o.work_order_parts.all())
        grand     = svcs_sum + parts_sum
        total_svcs  += svcs_sum
        total_parts_sum += parts_sum
        total_all   += grand

        client_name  = o.client_fio_static or (o.client_car.client.fio if o.client_car else '—')
        client_phone = (o.client_car.client.phone if o.client_car else '') or ''
        car_str      = o.car_details_static or '—'
        plate        = (o.client_car.license_plate if o.client_car else '') or ''

        ws1.append([
            o.completion_date, o.pk, client_name, client_phone,
            car_str, plate, float(svcs_sum), float(parts_sum), float(grand),
        ])
        style_row(ws1, i + 1, len(headers1), is_money_cols={7, 8, 9})
        ws1.cell(i + 2, 1).number_format = 'DD.MM.YYYY'

    total_row1 = len(orders) + 1
    add_total_row(ws1, total_row1, len(headers1), sum_cols={7, 8, 9})
    ws1.cell(total_row1 + 1, 7).value = float(total_svcs)
    ws1.cell(total_row1 + 1, 8).value = float(total_parts_sum)
    ws1.cell(total_row1 + 1, 9).value = float(total_all)

    # ══════════════════════════════════════════════════════════════════════
    # Лист 2: Услуги
    # ══════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet('Услуги')
    headers2 = [
        'Дата заказа', 'Дата завершения', '№ заказа', 'Статус заказа',
        'Клиент', 'Автомобиль', 'Тип услуги', 'Наименование услуги',
        'Нормо-часов', 'Коэффициент', 'Ставка (руб./ч)', 'Стоимость (руб.)',
    ]
    widths2 = [14, 16, 10, 18, 26, 24, 22, 30, 12, 12, 15, 16]
    style_header(ws2, headers2, widths2)

    # Завершённые — те же что в Сводке (completion_date), итог совпадёт точно
    wos_done = list(WorkOrderService.objects.select_related(
        'work_order', 'work_order__client_car__client', 'service__service_type',
    ).filter(
        work_order__status='Завершён',
        work_order__completion_date__gte=date_from,
        work_order__completion_date__lte=date_to,
    ).order_by('work_order__completion_date', 'work_order__pk'))
    # Незавершённые — по дате создания, отдельно для справки
    wos_pending = list(WorkOrderService.objects.select_related(
        'work_order', 'work_order__client_car__client', 'service__service_type',
    ).exclude(work_order__status='Завершён').filter(
        work_order__order_date__gte=date_from,
        work_order__order_date__lte=date_to,
    ).order_by('work_order__order_date', 'work_order__pk'))
    wos_qs = wos_done + wos_pending

    PENDING_FILL = PatternFill('solid', fgColor='FFFEF3C7')  # янтарный — незавершённые
    total_hours = Decimal('0'); total_svc_cost = Decimal('0')
    for i, wos in enumerate(wos_qs):
        o       = wos.work_order
        is_done = (o.status == 'Завершён')
        cname   = o.client_fio_static or (o.client_car.client.fio if o.client_car else '—')
        car     = o.car_details_static or '—'
        stype   = (wos.service.service_type.name if wos.service and wos.service.service_type else '—')
        sname   = wos.service_name_snapshot or (wos.service.name if wos.service else '—')
        rate    = wos.hourly_rate_snapshot or Decimal('0')
        price   = wos.final_price or Decimal('0')
        if is_done:
            total_hours    += wos.hours_applied
            total_svc_cost += price

        ws2.append([
            o.order_date, o.completion_date or '', o.pk, o.status,
            cname, car, stype, sname,
            float(wos.hours_applied), float(wos.complexity_factor),
            float(rate), float(price) if is_done else '',
        ])
        style_row(ws2, i + 1, len(headers2), is_money_cols={11, 12})
        if not is_done:
            for col in range(1, len(headers2) + 1):
                ws2.cell(i + 2, col).fill = PENDING_FILL
        ws2.cell(i + 2, 1).number_format = 'DD.MM.YYYY'
        if o.completion_date:
            ws2.cell(i + 2, 2).number_format = 'DD.MM.YYYY'

    tr2 = len(list(wos_qs)) + 1
    add_total_row(ws2, tr2, len(headers2), sum_cols={9, 11, 12})
    ws2.cell(tr2 + 1, 9).value  = float(total_hours)
    ws2.cell(tr2 + 1, 12).value = float(total_svc_cost)
    note2_row = tr2 + 2
    ws2.append(['ℹ Жёлтые строки — незавершённые заказы (в итог не включены). Итог по завершённым совпадает со Сводкой.'])
    ws2.merge_cells(start_row=note2_row, start_column=1, end_row=note2_row, end_column=len(headers2))
    _nc2 = ws2.cell(note2_row, 1)
    _nc2.font = Font(italic=True, name='Calibri', size=9, color='FF92400E')
    _nc2.fill = PatternFill('solid', fgColor=WARN_BG)
    _nc2.alignment = Alignment(horizontal='left', vertical='center')
    ws2.row_dimensions[note2_row].height = 22

    # ══════════════════════════════════════════════════════════════════════
    # Лист 3: Запасные части
    # ══════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet('Запасные части')
    headers3 = [
        'Дата заказа', 'Дата завершения', '№ заказа', 'Статус заказа',
        'Клиент', 'Артикул', 'Наименование детали', 'Бренд', 'Категория',
        'Кол-во (шт.)', 'Цена реализации (руб.)', 'Наценка (%)', 'Сумма (руб.)',
    ]
    widths3 = [14, 16, 10, 18, 26, 14, 30, 16, 18, 12, 22, 12, 15]
    style_header(ws3, headers3, widths3)

    # Завершённые — те же что в Сводке (completion_date)
    wop_done = list(WorkOrderPart.objects.select_related(
        'work_order', 'work_order__client_car__client', 'part',
    ).filter(
        work_order__status='Завершён',
        work_order__completion_date__gte=date_from,
        work_order__completion_date__lte=date_to,
    ).order_by('work_order__completion_date', 'work_order__pk'))
    # Незавершённые — по дате создания, для справки
    wop_pending = list(WorkOrderPart.objects.select_related(
        'work_order', 'work_order__client_car__client', 'part',
    ).exclude(work_order__status='Завершён').filter(
        work_order__order_date__gte=date_from,
        work_order__order_date__lte=date_to,
    ).order_by('work_order__order_date', 'work_order__pk'))
    wop_qs = wop_done + wop_pending

    total_parts_amt = Decimal('0')
    for i, wop in enumerate(wop_qs):
        o       = wop.work_order
        is_done = (o.status == 'Завершён')
        cname   = o.client_fio_static or (o.client_car.client.fio if o.client_car else '—')
        total   = wop.sale_price or Decimal('0')
        unit_p  = (total / wop.quantity).quantize(Decimal('0.01')) if wop.quantity else Decimal('0')
        if is_done:
            total_parts_amt += total

        ws3.append([
            o.order_date, o.completion_date or '', o.pk, o.status,
            cname,
            wop.part_article_snapshot or (wop.part.article if wop.part else '—'),
            wop.part_name_snapshot or (wop.part.name if wop.part else '—'),
            (wop.part.brand if wop.part else '—') or '—',
            (wop.part.category if wop.part else '—') or '—',
            wop.quantity, float(unit_p), float(wop.markup), float(total) if is_done else '',
        ])
        style_row(ws3, i + 1, len(headers3), is_money_cols={11, 13})
        if not is_done:
            for col in range(1, len(headers3) + 1):
                ws3.cell(i + 2, col).fill = PENDING_FILL
        ws3.cell(i + 2, 1).number_format = 'DD.MM.YYYY'
        if o.completion_date:
            ws3.cell(i + 2, 2).number_format = 'DD.MM.YYYY'

    tr3 = len(list(wop_qs)) + 1
    add_total_row(ws3, tr3, len(headers3), sum_cols={10, 11, 13})
    ws3.cell(tr3 + 1, 10).value = sum(wop.quantity for wop in wop_qs)
    ws3.cell(tr3 + 1, 13).value = float(total_parts_amt)
    note3_row = tr3 + 2
    ws3.append(['ℹ Жёлтые строки — незавершённые заказы (в итог не включены). Итог по завершённым совпадает со Сводкой.'])
    ws3.merge_cells(start_row=note3_row, start_column=1, end_row=note3_row, end_column=len(headers3))
    _nc3 = ws3.cell(note3_row, 1)
    _nc3.font = Font(italic=True, name='Calibri', size=9, color='FF92400E')
    _nc3.fill = PatternFill('solid', fgColor=WARN_BG)
    _nc3.alignment = Alignment(horizontal='left', vertical='center')
    ws3.row_dimensions[note3_row].height = 22

    # ══════════════════════════════════════════════════════════════════════
    # Лист 4: Поступление товаров
    # ══════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet('Поступление товаров')
    headers4 = [
        'Дата приёмки', '№ документа', 'Поставщик',
        'Артикул', 'Наименование', 'Бренд', 'Категория',
        'Место хранения', 'Кол-во (шт.)', 'Шт. в упаковке',
        'Цена за упаковку (руб.)', 'Цена за шт. (руб.)', 'Сумма (руб.)',
    ]
    widths4 = [14, 12, 22, 14, 30, 16, 18, 16, 12, 14, 22, 20, 15]
    style_header(ws4, headers4, widths4)

    si_qs = SupplyItem.objects.select_related(
        'document', 'document__supplier', 'part', 'location',
    ).filter(
        document__created_at__date__gte=date_from,
        document__created_at__date__lte=date_to,
    ).order_by('document__created_at', 'document__pk')

    total_supply = Decimal('0')
    for i, si in enumerate(si_qs):
        doc      = si.document
        supplier = doc.supplier.name if doc.supplier else '—'
        loc      = f'{si.location.rack}-{si.location.shelf}-{si.location.cell}'
        cpu      = si.price_per_unit
        row_sum  = cpu * si.quantity
        total_supply += row_sum

        ws4.append([
            doc.created_at.date(), doc.pk, supplier,
            si.part.article, si.part.name,
            si.part.brand or '—', si.part.category or '—',
            loc, si.quantity, si.pkg_qty,
            float(si.purchase_price), float(cpu), float(row_sum),
        ])
        style_row(ws4, i + 1, len(headers4), is_money_cols={11, 12, 13})
        ws4.cell(i + 2, 1).number_format = 'DD.MM.YYYY'

    tr4 = len(list(si_qs)) + 1
    add_total_row(ws4, tr4, len(headers4), sum_cols={9, 11, 12, 13})
    ws4.cell(tr4 + 1, 9).value  = sum(si.quantity for si in si_qs)
    ws4.cell(tr4 + 1, 13).value = float(total_supply)

    # ══════════════════════════════════════════════════════════════════════
    # Лист 5: Фонд оплаты труда
    # Заработок = часы (доля) × нормо-час × К.ЗП
    # К. сложности — это надбавка сервиса к цене клиента, не к зарплате
    # ══════════════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet('Фонд оплаты труда')
    ws5.append([f'Период: {period_label}   |   Формула зарплаты: Часы (доля) × Нормо-час × К.ЗП сотрудника'])
    ws5.merge_cells('A1:K1')
    note5 = ws5['A1']
    note5.font = Font(italic=True, name='Calibri', size=9, color='FF334155')
    note5.alignment = Alignment(horizontal='left', vertical='center')
    ws5.row_dimensions[1].height = 18

    headers5 = [
        'Дата завершения', '№ заказа', 'Клиент', 'Автомобиль',
        'Сотрудник', 'Должность', 'Услуга',
        'Часов (доля)', 'Нормо-час (руб./ч)', 'К. ЗП', 'Заработок (руб.)',
    ]
    widths5 = [16, 10, 26, 24, 28, 18, 30, 13, 18, 10, 16]
    ws5.append(headers5)
    for col, w in enumerate(widths5, 1):
        cell = ws5.cell(2, col)
        cell.fill   = hdr_fill()
        cell.font   = hdr_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        ws5.column_dimensions[get_column_letter(col)].width = w
    ws5.row_dimensions[2].height = 30
    ws5.freeze_panes = 'A3'

    emp_qs = WorkOrderServiceEmployee.objects.select_related(
        'employee',
        'work_order_service__work_order',
        'work_order_service__work_order__client_car__client',
        'work_order_service',
    ).filter(
        work_order_service__work_order__status='Завершён',
        work_order_service__work_order__completion_date__gte=date_from,
        work_order_service__work_order__completion_date__lte=date_to,
    ).order_by(
        'work_order_service__work_order__completion_date',
        'work_order_service__work_order__pk',
    )

    total_fot = Decimal('0')
    total_fot_hours = Decimal('0')
    for i, asgn in enumerate(emp_qs):
        wos   = asgn.work_order_service
        o     = wos.work_order
        emp   = asgn.employee
        n     = wos.assignments.count() or 1
        rate  = wos.hourly_rate_snapshot or Decimal('0')
        coeff = asgn.salary_coefficient_snapshot if asgn.salary_coefficient_snapshot is not None \
            else ((emp.salary_coefficient if emp else None) or Decimal('1'))
        hours_share = (wos.hours_applied / Decimal(n)).quantize(Decimal('0.01'))
        earn  = (hours_share * rate * coeff).quantize(Decimal('0.01'))
        total_fot += earn
        total_fot_hours += hours_share
        cname = o.client_fio_static or (o.client_car.client.fio if o.client_car else '—')
        car   = o.car_details_static or '—'
        sname = wos.service_name_snapshot or (wos.service.name if wos.service else '—')
        emp_name = asgn.employee_name_snapshot or (emp.name if emp else '—')
        emp_pos  = (emp.position if emp else None) or '—'

        row_num = i + 3
        ws5.append([
            o.completion_date, o.pk, cname, car,
            emp_name, emp_pos, sname,
            float(hours_share), float(rate), float(coeff), float(earn),
        ])
        fill = alt_fill(i)
        for col in range(1, len(headers5) + 1):
            cell = ws5.cell(row_num, col)
            cell.fill   = fill
            cell.font   = data_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')
            if col in {9, 11}:
                cell.number_format = '#,##0.00'
        ws5.cell(row_num, 1).number_format = 'DD.MM.YYYY'

    tot5_row = len(list(emp_qs)) + 3
    ws5.append(['ИТОГО', '', '', '', '', '', '', float(total_fot_hours), '', '', float(total_fot)])
    for col in range(1, len(headers5) + 1):
        cell = ws5.cell(tot5_row, col)
        cell.fill   = tot_fill()
        cell.font   = bold_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center')
        if col in {8, 11}:
            cell.number_format = '#,##0.00'

    # ══════════════════════════════════════════════════════════════════════
    # Лист 6: Сводка — выручка, затраты, прибыль сервиса
    # ══════════════════════════════════════════════════════════════════════
    ws6 = wb.create_sheet('Сводка')

    GREEN_BG  = 'FFD1FAE5'
    GREEN_HDR = 'FF065F46'
    RED_BG    = 'FFFEE2E2'
    RED_HDR   = 'FF991B1B'
    GRAY_BG   = 'FFF1F5F9'

    def _ws6_section_hdr(ws, label, bg, fg, n_cols=3):
        ws.append([label])
        r = ws.max_row
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)
        c = ws.cell(r, 1)
        c.fill = PatternFill('solid', fgColor=bg)
        c.font = Font(bold=True, name='Calibri', size=10, color=fg)
        c.alignment = Alignment(horizontal='left', vertical='center')
        c.border = thin_border
        ws.row_dimensions[r].height = 20

    def _ws6_row(ws, label, value, note='', bold=False, bg=None):
        ws.append([label, value, note])
        r = ws.max_row
        for col in range(1, 4):
            c = ws.cell(r, col)
            c.font = bold_font if bold else data_font
            c.border = thin_border
            c.alignment = Alignment(vertical='center',
                                    horizontal='right' if col == 2 else 'left')
            if bg:
                c.fill = PatternFill('solid', fgColor=bg)
            if col == 2:
                c.number_format = '#,##0.00'

    def _ws6_blank(ws):
        ws.append(['', '', ''])
        ws.row_dimensions[ws.max_row].height = 6

    # Заголовок листа
    ws6.append([f'Сводка доходов и расходов сервиса  |  {period_label}'])
    ws6.merge_cells('A1:C1')
    title6 = ws6['A1']
    title6.font = Font(bold=True, name='Calibri', size=12, color='FF1E293B')
    title6.alignment = Alignment(horizontal='left', vertical='center')
    ws6.row_dimensions[1].height = 24
    ws6.append(['ℹ Сводка — только ЗАВЕРШЁННЫЕ заказы, фильтр по дате завершения. Листы «Услуги» и «Запчасти» показывают все заказы (жёлтые = незавершённые, не входят в итог).', '', ''])
    ws6.merge_cells('A2:C2')
    _note_ws6 = ws6.cell(2, 1)
    _note_ws6.font = Font(italic=True, name='Calibri', size=9, color='FF92400E')
    _note_ws6.fill = PatternFill('solid', fgColor=WARN_BG)
    _note_ws6.alignment = Alignment(horizontal='left', vertical='center')
    ws6.row_dimensions[2].height = 22

    ws6.append(['Показатель', 'Сумма (руб.)', 'Примечание'])
    for col in range(1, 4):
        c = ws6.cell(3, col)
        c.fill = hdr_fill()
        c.font = hdr_font
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin_border
    ws6.row_dimensions[3].height = 22
    ws6.column_dimensions['A'].width = 42
    ws6.column_dimensions['B'].width = 18
    ws6.column_dimensions['C'].width = 40
    ws6.freeze_panes = 'A4'

    _ws6_blank(ws6)

    # ── Выручка ──
    _ws6_section_hdr(ws6, '▸ ВЫРУЧКА', 'FF1E3A5F', 'FFFFFFFF')
    # Услуги: выручка = hours × rate × complexity (цена для клиента)
    svc_revenue = sum(
        w.final_price or Decimal('0')
        for o in orders
        for w in o.work_order_services.all()
    )
    # Запчасти: sale_price уже содержит итог (цена × кол-во), умножать на qty не нужно
    parts_revenue = sum(
        (w.sale_price or Decimal('0'))
        for o in orders
        for w in o.work_order_parts.all()
    )
    total_revenue = svc_revenue + parts_revenue
    _ws6_row(ws6, 'Выручка от услуг', float(svc_revenue),
             'Нормо-часы × ставка × коэф. сложности (цена для клиента)')
    _ws6_row(ws6, 'Выручка от запасных частей', float(parts_revenue),
             'Цена реализации × количество')
    _ws6_row(ws6, 'ИТОГО ВЫРУЧКА', float(total_revenue), '', bold=True, bg='FFE0F2FE')

    _ws6_blank(ws6)

    # ── Расходы ──
    _ws6_section_hdr(ws6, '▸ РАСХОДЫ', RED_HDR, 'FFFFFFFF')
    # ФОТ по услугам завершённых заказов за период
    _ws6_row(ws6, 'Фонд оплаты труда (ФОТ)', float(total_fot),
             'Часы (доля) × нормо-час × К.ЗП — см. лист «Фонд оплаты труда»')
    # Себестоимость запчастей: sale / (1 + markup/100) × qty
    parts_cost = Decimal('0')
    for o in orders:
        for w in o.work_order_parts.all():
            if w.sale_price and w.markup is not None:
                parts_cost += w.sale_price / (1 + w.markup / Decimal('100'))
            elif w.sale_price:
                parts_cost += w.sale_price
    parts_cost = parts_cost.quantize(Decimal('0.01'))
    _ws6_row(ws6, 'Себестоимость запасных частей', float(parts_cost),
             'Цена продажи ÷ (1 + наценка%) × количество')
    total_costs = total_fot + parts_cost
    _ws6_row(ws6, 'ИТОГО РАСХОДЫ', float(total_costs), '', bold=True, bg='FFFEE2E2')

    _ws6_blank(ws6)

    # ── Прибыль ──
    _ws6_section_hdr(ws6, '▸ ПРИБЫЛЬ', GREEN_HDR, 'FFFFFFFF')
    svc_profit = svc_revenue - total_fot
    parts_profit = parts_revenue - parts_cost
    total_profit = total_revenue - total_costs
    _ws6_row(ws6, 'Прибыль от услуг', float(svc_profit),
             'Выручка от услуг − ФОТ (включает надбавку за сложность)')
    _ws6_row(ws6, 'Прибыль от запасных частей', float(parts_profit),
             'Выручка от запчастей − себестоимость (наценка)')
    profit_bg = 'FFD1FAE5' if total_profit >= 0 else 'FFFEE2E2'
    _ws6_row(ws6, 'ИТОГО ПРИБЫЛЬ', float(total_profit), '', bold=True, bg=profit_bg)

    _ws6_blank(ws6)

    # ── Справочно ──
    _ws6_section_hdr(ws6, '▸ СПРАВОЧНО', 'FF475569', 'FFFFFFFF')
    _ws6_row(ws6, 'Поступление товаров (закупка)', float(total_supply),
             'Все приходы за период — см. лист «Поступление товаров»')
    _ws6_row(ws6, 'Завершённых заказов', len(orders), '')
    margin_pct = (total_profit / total_revenue * 100).quantize(Decimal('0.1')) if total_revenue else Decimal('0')
    _ws6_row(ws6, 'Рентабельность', f'{margin_pct}%',
             'Прибыль ÷ Выручка × 100%')

    # ── Генерация файла ──────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f'accounting_{date_from.strftime("%Y%m%d")}_{date_to.strftime("%Y%m%d")}.xlsx'
    resp = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


@login_required  # Требует авторизацию
@staff_required
def export_db(request):
    buffer = StringIO()
    call_command('dumpdata', stdout=buffer, natural_foreign=True, natural_primary=True)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type='application/json')
    response['Content-Disposition'] = 'attachment; filename="full_db_export.json"'
    return response


@login_required
@staff_required
def import_db(request):
    if request.method == 'POST':
        if 'db_file' in request.FILES:
            db_file = request.FILES['db_file']
            try:
                # Читаем содержимое файла как байты
                raw_data = db_file.read()

                # Определяем кодировку
                result = chardet.detect(raw_data)
                encoding = result['encoding']
                confidence = result['confidence']

                if encoding is None or confidence < 0.5:
                    messages.error(request, 'Не удалось определить кодировку файла.')
                    return redirect('orders_list')

                # Декодируем данные с определённой кодировкой
                try:
                    json_data = raw_data.decode(encoding)
                except UnicodeDecodeError:
                    # Попробуем Windows-1251 как запасной вариант
                    try:
                        json_data = raw_data.decode('windows-1251')
                    except UnicodeDecodeError as e:
                        messages.error(request, f'Ошибка декодирования файла: {str(e)}')
                        return redirect('orders_list')

                # Очищаем базу данных перед загрузкой
                with transaction.atomic():
                    all_models = [model for model in apps.get_models() if model._meta.app_label != 'contenttypes']
                    for model in all_models:
                        if model._meta.db_table != 'django_migrations':
                            model.objects.all().delete()

                # Создаём временный файл в UTF-8
                with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False, encoding='utf-8') as temp_file:
                    temp_file.write(json_data)
                    temp_file_path = temp_file.name

                try:
                    # Загружаем данные из временного файла
                    call_command('loaddata', temp_file_path, format='json', verbosity=0)
                    messages.success(request, 'База данных успешно импортирована и полностью перезаписана!')
                finally:
                    # Удаляем временный файл
                    os.remove(temp_file_path)

            except Exception as e:
                messages.error(request, f'Ошибка при импорте: {str(e)}')
            return redirect('orders_list')
    return render(request, 'import_db.html', {'cancel_url': request.META.get('HTTP_REFERER', '/')})



@login_required
def order_mechanic_pdf(request, pk):
    """Generate a detailed PDF work order for the mechanic."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle,
        Paragraph, Spacer, HRFlowable,
    )
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import io, os
    from warehouse.models import WorkOrderService, WorkOrderPart

    order = get_object_or_404(Order, pk=pk)

    # Register Cyrillic font
    _fonts = [
        '/Library/Fonts/Arial Unicode.ttf',
        '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf',
        '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
    ]
    fn = 'Helvetica'
    for _fp in _fonts:
        if os.path.exists(_fp):
            try:
                pdfmetrics.registerFont(TTFont('_MechFont', _fp))
                fn = '_MechFont'
            except Exception:
                pass
            break

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    BLUE_HDR  = colors.HexColor('#1e40af')
    STRIPE    = colors.HexColor('#f0f4ff')
    GREY_RULE = colors.HexColor('#cbd5e1')

    H1   = ParagraphStyle('H1',   fontName=fn, fontSize=14, spaceAfter=4,  leading=18, textColor=colors.HexColor('#1e3a5f'))
    H2   = ParagraphStyle('H2',   fontName=fn, fontSize=11, spaceAfter=3,  leading=14, textColor=BLUE_HDR)
    NR   = ParagraphStyle('NR',   fontName=fn, fontSize=9,  spaceAfter=2,  leading=12)
    SM   = ParagraphStyle('SM',   fontName=fn, fontSize=8,  spaceAfter=1,  leading=11, textColor=colors.HexColor('#6b7280'))
    # Cell styles — Paragraph inside a table cell wraps text automatically
    CH   = ParagraphStyle('CH',   fontName=fn, fontSize=8,  leading=10, textColor=colors.white)
    CB   = ParagraphStyle('CB',   fontName=fn, fontSize=8,  leading=10)

    def p(text, style=None):
        """Wrap a value in Paragraph so it word-wraps inside table cells."""
        return Paragraph(str(text) if text is not None else '—', style or CB)

    def ph(text):
        """Header cell — white text."""
        return Paragraph(str(text), CH)

    def tbl(data, col_widths, header_rows=1):
        t = Table(data, colWidths=col_widths, repeatRows=header_rows)
        t.setStyle(TableStyle([
            ('BACKGROUND',    (0, 0), (-1, header_rows - 1), BLUE_HDR),
            ('GRID',          (0, 0), (-1, -1), 0.3, GREY_RULE),
            ('ROWBACKGROUNDS',(0, header_rows), (-1, -1), [colors.white, STRIPE]),
            ('VALIGN',        (0, 0), (-1, -1), 'TOP'),
            ('TOPPADDING',    (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING',   (0, 0), (-1, -1), 5),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 4),
        ]))
        return t

    story = []

    # ── Title ────────────────────────────────────
    story.append(Paragraph(f'Заказ-наряд №{order.id}', H1))
    story.append(Paragraph(f'Дата: {order.order_date.strftime("%d.%m.%Y")}   |   Статус: {order.status}', SM))
    story.append(HRFlowable(width='100%', thickness=1, color=BLUE_HDR, spaceAfter=8))

    # ── Client & Car ─────────────────────────────
    story.append(Paragraph('Клиент и автомобиль', H2))
    client_name = order.client_fio_static or (order.client_car.client.fio if order.client_car else '—')
    client_phone = ''
    if order.client_car and order.client_car.client.phone:
        client_phone = order.client_car.client.phone

    car_str = order.car_details_static or '—'
    car_extra = ''
    if order.client_car:
        cc = order.client_car
        parts_extra = []
        if cc.vin:   parts_extra.append(f'VIN: {cc.vin}')
        if cc.year:  parts_extra.append(f'Год: {cc.year}')
        if cc.color: parts_extra.append(f'Цвет: {cc.color}')
        car_extra = '   '.join(parts_extra)

    LBL = ParagraphStyle('LBL', fontName=fn, fontSize=8, leading=10, textColor=colors.HexColor('#374151'))
    info_data = [[ph('Параметр'), ph('Значение')]]
    info_data.append([p('Клиент', LBL), p(client_name)])
    if client_phone:
        info_data.append([p('Телефон', LBL), p(client_phone)])
    info_data.append([p('Автомобиль', LBL), p(car_str)])
    if car_extra:
        info_data.append([p('Доп. данные', LBL), p(car_extra)])
    if order.comment:
        info_data.append([p('Описание', LBL), p(order.comment)])
    story.append(tbl(info_data, [3.5*cm, 13.7*cm]))
    story.append(Spacer(1, 0.4*cm))

    # ── Services ─────────────────────────────────
    wos_list = list(
        order.work_order_services
        .prefetch_related('assignments__employee')
        .all()
    )
    if wos_list:
        story.append(Paragraph('Услуги', H2))
        # Widths: Услуга(4.4) Часы(1.3) Коэф.(1.1) Нормо-час(2.0) Исполнители(8.4) = 17.2
        svc_data = [[ph('Услуга'), ph('Часы'), ph('Коэф.'), ph('Нормо-час'), ph('Исполнители')]]
        for wos in wos_list:
            name  = wos.service_name_snapshot or (wos.service.name if wos.service else '—')
            rate  = str(wos.hourly_rate_snapshot or '—')
            emps  = ', '.join(
                a.employee_name_snapshot or (a.employee.name if a.employee else '—')
                for a in wos.assignments.all()
            ) or '—'
            svc_data.append([
                p(name), p(wos.hours_applied), p(wos.complexity_factor),
                p(rate), p(emps),
            ])
        story.append(tbl(svc_data, [4.4*cm, 1.3*cm, 1.1*cm, 2.0*cm, 8.4*cm]))
        story.append(Spacer(1, 0.4*cm))

    # ── Parts ────────────────────────────────────
    wop_list = list(
        order.work_order_parts
        .select_related('part', 'work_order_service')
        .prefetch_related('part__stock_entries__location')
        .all()
    )
    if wop_list:
        story.append(Paragraph('Запчасти', H2))
        # Widths: Артикул(2.1) Название(5.5) Кол-во(1.3) Место(2.8) Статус(2.2) Услуга(3.3) = 17.2
        part_data = [[ph('Артикул'), ph('Название'), ph('Кол-во'), ph('Место хранения'), ph('Статус'), ph('Услуга')]]
        for wop in wop_list:
            entry = wop.part.stock_entries.first() if wop.part else None
            loc   = entry.location.label if entry else '—'
            svc_name = '—'
            if wop.work_order_service:
                svc_name = (wop.work_order_service.service_name_snapshot or
                            (wop.work_order_service.service.name if wop.work_order_service.service else '—'))
            part_data.append([
                p(wop.part_article_snapshot or (wop.part.article if wop.part else '—')),
                p(wop.part_name_snapshot or (wop.part.name if wop.part else '—')),
                p(wop.quantity),
                p(loc),
                p(wop.get_status_display()),
                p(svc_name),
            ])
        story.append(tbl(part_data, [2.1*cm, 5.5*cm, 1.3*cm, 2.8*cm, 2.2*cm, 3.3*cm]))
        story.append(Spacer(1, 0.4*cm))

    doc.build(story)
    buf.seek(0)
    resp = HttpResponse(buf, content_type='application/pdf')
    resp['Content-Disposition'] = f'inline; filename="order_{order.pk}_mechanic.pdf"'
    return resp


def _make_customer_pdf(order, is_final: bool):
    """Shared builder for customer-facing work order PDFs."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle,
        Paragraph, Spacer, HRFlowable, KeepTogether,
    )
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from decimal import Decimal
    import io, os
    from warehouse.models import WorkOrderService, WorkOrderPart

    _fonts = [
        '/Library/Fonts/Arial Unicode.ttf',
        '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf',
        '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
    ]
    fn = 'Helvetica'
    for _fp in _fonts:
        if os.path.exists(_fp):
            try:
                pdfmetrics.registerFont(TTFont('_CustFont', _fp))
                fn = '_CustFont'
            except Exception:
                pass
            break

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    BLUE      = colors.HexColor('#1e40af')
    BLUE_DARK = colors.HexColor('#1e3a8a')
    GREY      = colors.HexColor('#cbd5e1')
    STRIPE    = colors.HexColor('#f0f4ff')
    DARKTEXT  = colors.HexColor('#1e293b')
    MUTED     = colors.HexColor('#64748b')
    WARN_BG   = colors.HexColor('#fef9c3')
    WARN_TXT  = colors.HexColor('#92400e')

    H1  = ParagraphStyle('H1',  fontName=fn, fontSize=16, leading=20, textColor=BLUE_DARK, spaceAfter=2)
    H1S = ParagraphStyle('H1S', fontName=fn, fontSize=10, leading=13, textColor=MUTED, spaceAfter=4)
    SUB = ParagraphStyle('SUB', fontName=fn, fontSize=8,  leading=11, textColor=MUTED, spaceAfter=4)
    H2  = ParagraphStyle('H2',  fontName=fn, fontSize=10, leading=13, textColor=BLUE, spaceBefore=8, spaceAfter=4)
    NR  = ParagraphStyle('NR',  fontName=fn, fontSize=9,  leading=12, textColor=DARKTEXT)
    SIG = ParagraphStyle('SIG', fontName=fn, fontSize=9,  leading=14, textColor=DARKTEXT)
    PRE = ParagraphStyle('PRE', fontName=fn, fontSize=8,  leading=11,
                         textColor=WARN_TXT, backColor=WARN_BG,
                         spaceAfter=8, leftIndent=6, rightIndent=6)
    ORG = ParagraphStyle('ORG', fontName=fn, fontSize=8,  leading=11, textColor=DARKTEXT)
    ORG_LBL = ParagraphStyle('ORGL', fontName=fn, fontSize=7, leading=9, textColor=MUTED)
    LEGAL = ParagraphStyle('LEG', fontName=fn, fontSize=7, leading=10,
                           textColor=MUTED, spaceAfter=6)

    def ch(text):
        return Paragraph(str(text), ParagraphStyle('ch', fontName=fn, fontSize=8, leading=10,
                                                    textColor=colors.white))

    def cb(text, align='LEFT'):
        return Paragraph(str(text) if text is not None else '—',
                         ParagraphStyle('cb', fontName=fn, fontSize=8, leading=10, alignment={'RIGHT':2,'CENTER':1}.get(align, 0)))

    def cb_bold(text):
        return Paragraph(str(text) if text is not None else '—',
                         ParagraphStyle('cbb', fontName=fn, fontSize=9, leading=11, textColor=DARKTEXT))

    def make_table(data, col_widths):
        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND',    (0, 0), (-1, 0), BLUE),
            ('GRID',          (0, 0), (-1, -1), 0.3, GREY),
            ('ROWBACKGROUNDS',(0, 1), (-1, -1), [colors.white, STRIPE]),
            ('VALIGN',        (0, 0), (-1, -1), 'TOP'),
            ('TOPPADDING',    (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING',   (0, 0), (-1, -1), 5),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 4),
        ]))
        return t

    # ── Extract org data ─────────────────────────
    org = order.org_snapshot or {}
    org_type    = org.get('org_type', '')
    org_name    = org.get('org_name', '')
    inn         = org.get('inn', '')
    kpp         = org.get('kpp', '')
    ogrn        = org.get('ogrn', '')
    ogrnip      = org.get('ogrnip', '')
    org_address = org.get('org_address', '')
    org_phone   = org.get('org_phone', '')
    org_email   = org.get('org_email', '')
    warranty_days = org.get('warranty_days', 30)

    story = []
    PAGE_W = A4[0] - 4*cm  # usable width

    # ── Org header block ─────────────────────────
    if org_name or org_type:
        full_org_name = f'{org_type} {org_name}'.strip() if org_name else org_type
        org_detail_parts = []
        if inn:
            org_detail_parts.append(f'ИНН: {inn}')
        if kpp and org_type != 'ИП':
            org_detail_parts.append(f'КПП: {kpp}')
        reg_no = ogrnip if org_type == 'ИП' else ogrn
        if reg_no:
            label = 'ОГРНИП' if org_type == 'ИП' else 'ОГРН'
            org_detail_parts.append(f'{label}: {reg_no}')
        org_detail_str = '   |   '.join(org_detail_parts) if org_detail_parts else ''

        contact_parts = []
        if org_address:
            contact_parts.append(org_address)
        if org_phone:
            contact_parts.append(f'Тел.: {org_phone}')
        if org_email:
            contact_parts.append(f'Email: {org_email}')
        contact_str = '   |   '.join(contact_parts)

        org_data = [[
            Paragraph(full_org_name, ParagraphStyle('on', fontName=fn, fontSize=11, leading=14, textColor=BLUE_DARK)),
            Paragraph(
                f'<font size="7" color="#64748b">{org_detail_str}</font><br/>'
                f'<font size="7" color="#64748b">{contact_str}</font>',
                ParagraphStyle('od', fontName=fn, fontSize=7, leading=10, textColor=MUTED),
            ),
        ]]
        org_tbl = Table(org_data, colWidths=[7*cm, PAGE_W - 7*cm])
        org_tbl.setStyle(TableStyle([
            ('VALIGN',      (0,0), (-1,-1), 'MIDDLE'),
            ('LEFTPADDING', (0,0), (-1,-1), 0),
            ('RIGHTPADDING',(0,0), (-1,-1), 0),
            ('TOPPADDING',  (0,0), (-1,-1), 0),
            ('BOTTOMPADDING',(0,0), (-1,-1), 0),
        ]))
        story.append(org_tbl)
        story.append(HRFlowable(width='100%', thickness=1, color=BLUE, spaceBefore=6, spaceAfter=6))

    # ── Title ────────────────────────────────────
    if is_final:
        doc_title = f'АКТ ВЫПОЛНЕННЫХ РАБОТ  №{order.id}'
        doc_subtitle = 'Заказ-наряд'
    else:
        doc_title = f'ЗАКАЗ-НАРЯД  №{order.id}'
        doc_subtitle = 'Предварительный — цены могут измениться до выдачи автомобиля'

    story.append(Paragraph(doc_title, H1))
    story.append(Paragraph(doc_subtitle, H1S))

    date_parts = [f'Дата оформления: {order.order_date.strftime("%d.%m.%Y")}']
    if is_final and order.completion_date:
        date_parts.append(f'Дата выдачи: {order.completion_date.strftime("%d.%m.%Y")}')
    story.append(Paragraph('   |   '.join(date_parts), SUB))

    if not is_final:
        story.append(Paragraph(
            'Предварительный документ. Окончательная стоимость определяется после выполнения работ.',
            PRE
        ))

    story.append(HRFlowable(width='100%', thickness=0.5, color=GREY, spaceAfter=6))

    # ── Client + Car info (two-column) ────────────
    cs = order.car_snapshot or {}
    client_name  = cs.get('client_fio') or order.client_fio_static or (order.client_car.client.fio if order.client_car else '—')
    client_phone = cs.get('client_phone') or (order.client_car.client.phone if order.client_car else '') or ''
    vin_str      = cs.get('vin') or (order.client_car.vin if order.client_car else '') or ''
    plate_str    = cs.get('plate') or (order.client_car.license_plate if order.client_car else '') or ''
    car_make     = cs.get('make') or (order.client_car.make.name if order.client_car else '') or order.car_details_static or ''
    car_model    = cs.get('model') or (order.client_car.model.name if order.client_car else '') or ''
    car_year     = cs.get('year') or (order.client_car.year if order.client_car else None)
    car_color    = cs.get('color') or (order.client_car.color if order.client_car else '') or ''

    def info_block(rows):
        """rows = [(label, value), ...]"""
        items = []
        for lbl, val in rows:
            items.append(Paragraph(lbl, ORG_LBL))
            items.append(Paragraph(str(val) if val else '—', ORG))
        return items

    col_w = (PAGE_W - 0.5*cm) / 2

    client_block = info_block([
        ('Клиент', client_name),
        ('Телефон', client_phone or '—'),
    ])
    car_name = f'{car_make} {car_model}'.strip() or '—'
    if car_year:
        car_name += f' ({car_year} г.)'
    car_rows = [
        ('Марка / Модель', car_name),
        ('Госномер', plate_str or '—'),
        ('VIN', vin_str or '—'),
    ]
    if car_color:
        car_rows.append(('Цвет', car_color))
    if order.mileage is not None:
        if order.mileage_prev is not None:
            mileage_val = f'{order.mileage} (изменён с {order.mileage_prev})'
            if order.mileage_change_reason:
                mileage_val += f'; причина: {order.mileage_change_reason}'
            car_rows.append(('Пробег (км)', mileage_val))
        else:
            car_rows.append(('Пробег при приёмке, км', order.mileage))
    car_block = info_block(car_rows)

    two_col = Table(
        [[client_block, car_block]],
        colWidths=[col_w, col_w + 0.5*cm],
    )
    two_col.setStyle(TableStyle([
        ('VALIGN',          (0,0), (-1,-1), 'TOP'),
        ('LEFTPADDING',     (0,0), (-1,-1), 0),
        ('RIGHTPADDING',    (0,0), (-1,-1), 4),
        ('TOPPADDING',      (0,0), (-1,-1), 0),
        ('BOTTOMPADDING',   (0,0), (-1,-1), 0),
    ]))
    story.append(two_col)

    if order.comment:
        story.append(Spacer(1, 0.2*cm))
        story.append(Paragraph(f'Описание: {order.comment}', ORG))

    story.append(Spacer(1, 0.4*cm))

    # ── Services ─────────────────────────────────
    wos_list = list(order.work_order_services.all())
    wop_list = list(order.work_order_parts.select_related('part').all())

    if wos_list:
        story.append(Paragraph('Выполненные работы', H2))
        svc_rows = [[ch('№'), ch('Наименование работы'), ch('Часы'), ch('Коэф.'), ch('Сумма, руб.')]]
        for i, wos in enumerate(wos_list, 1):
            name  = wos.service_name_snapshot or (wos.service.name if wos.service else '—')
            price = f'{wos.final_price:.2f}' if wos.final_price else '—'
            svc_rows.append([cb(i), cb(name), cb(wos.hours_applied), cb(wos.complexity_factor), cb(price, 'RIGHT')])
        svc_tbl = make_table(svc_rows, [0.8*cm, 10.5*cm, 1.5*cm, 1.5*cm, 3.0*cm])
        svc_tbl.setStyle(TableStyle([
            ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ]))
        story.append(svc_tbl)
        story.append(Spacer(1, 0.3*cm))

    # ── Parts ────────────────────────────────────
    if wop_list:
        story.append(Paragraph('Запчасти и материалы', H2))
        part_rows = [[ch('№'), ch('Артикул'), ch('Наименование'), ch('Кол-во'), ch('Цена, руб.'), ch('Сумма, руб.')]]
        for i, wop in enumerate(wop_list, 1):
            unit_price = f'{wop.sale_price / wop.quantity:.2f}' if wop.sale_price and wop.quantity else '—'
            total      = f'{wop.sale_price:.2f}' if wop.sale_price else '—'
            article = wop.part_article_snapshot or (wop.part.article if wop.part else '—')
            name    = wop.part_name_snapshot    or (wop.part.name    if wop.part else '—')
            part_rows.append([cb(i), cb(article), cb(name), cb(wop.quantity, 'RIGHT'),
                               cb(unit_price, 'RIGHT'), cb(total, 'RIGHT')])
        part_tbl = make_table(part_rows, [0.8*cm, 2.2*cm, 7.2*cm, 1.3*cm, 2.4*cm, 3.4*cm])
        part_tbl.setStyle(TableStyle([
            ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ]))
        story.append(part_tbl)
        story.append(Spacer(1, 0.3*cm))

    # ── Totals ───────────────────────────────────
    svc_total  = sum(w.final_price  or Decimal('0') for w in wos_list)
    part_total = sum(w.sale_price   or Decimal('0') for w in wop_list)
    grand      = svc_total + part_total

    BOLD_NR = ParagraphStyle('BNR', fontName=fn, fontSize=10, leading=13, textColor=DARKTEXT)
    total_rows = [
        [Paragraph('', NR), Paragraph('Работы:', NR), Paragraph(f'{svc_total:.2f} руб.', NR)],
        [Paragraph('', NR), Paragraph('Запчасти и материалы:', NR), Paragraph(f'{part_total:.2f} руб.', NR)],
        [Paragraph('', BOLD_NR), Paragraph('ИТОГО:', BOLD_NR), Paragraph(f'{grand:.2f} руб.', BOLD_NR)],
    ]
    if order.payment_method:
        labels = dict(Order.PAYMENT_METHOD_CHOICES)
        total_rows.append([
            Paragraph('', NR),
            Paragraph('Форма оплаты:', NR),
            Paragraph(labels.get(order.payment_method, order.payment_method), NR),
        ])
    tot_tbl = Table(total_rows, colWidths=[9.7*cm, 4.3*cm, 3.3*cm])
    tot_tbl.setStyle(TableStyle([
        ('ALIGN',        (1, 0), (-1, -1), 'RIGHT'),
        ('LINEABOVE',    (0, 2), (-1, 2), 0.8, DARKTEXT),
        ('LINEBELOW',    (0, 2), (-1, 2), 0.8, DARKTEXT),
        ('TOPPADDING',   (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING',(0, 0), (-1, -1), 3),
        ('FONTNAME',     (0, 0), (-1, -1), fn),
        ('FONTSIZE',     (0, 0), (-1, 1), 8),
        ('FONTSIZE',     (0, 2), (-1, 2), 10),
        ('FONTSIZE',     (0, 3), (-1, 3), 8),
    ]))
    story.append(tot_tbl)

    # ── Warranty block ───────────────────────────
    if is_final and warranty_days:
        story.append(Spacer(1, 0.4*cm))
        story.append(HRFlowable(width='100%', thickness=0.5, color=GREY, spaceAfter=5))
        story.append(Paragraph(
            f'Гарантийные обязательства: гарантийный срок на выполненные работы составляет '
            f'{warranty_days} ({"тридцать" if warranty_days == 30 else str(warranty_days)}) календарных дней '
            f'с даты подписания настоящего акта.',
            LEGAL
        ))

    # ── Signatures ───────────────────────────────
    story.append(Spacer(1, 0.6*cm))
    story.append(HRFlowable(width='100%', thickness=0.5, color=GREY, spaceAfter=8))

    HINT = ParagraphStyle('hint', fontName=fn, fontSize=7,
                          leading=9, textColor=colors.HexColor('#9ca3af'))
    date_str = (order.completion_date.strftime('%d.%m.%Y')
                if is_final and order.completion_date else '________________')

    # Collect unique employees
    seen_emp_ids = set()
    unique_employees = []
    for wos in order.work_order_services.prefetch_related('assignments__employee').all():
        for asgn in wos.assignments.all():
            emp = asgn.employee
            if emp.pk not in seen_emp_ids:
                seen_emp_ids.add(emp.pk)
                unique_employees.append(emp)
    if not unique_employees:
        unique_employees = [None]

    sig_data = []
    for emp in unique_employees:
        fio_str = emp.name if emp else ''
        sig_data.append([
            Paragraph('Исполнитель:', SIG),
            Paragraph('_______________________', SIG),
            Paragraph(f'/ {fio_str or "______________________"} /', SIG),
            Paragraph(f'Дата: {date_str}', SIG),
        ])
        sig_data.append([
            Paragraph('', SIG),
            Paragraph('(подпись)', HINT),
            Paragraph('(расшифровка подписи)', HINT),
            Paragraph('', SIG),
        ])
        sig_data.append([Paragraph('', SIG)] * 4)

    cs2 = order.car_snapshot or {}
    client_fio = cs2.get('client_fio') or order.client_fio_static or (order.client_car.client.fio if order.client_car else '')
    sig_data += [
        [
            Paragraph('Клиент:', SIG),
            Paragraph('_______________________', SIG),
            Paragraph(f'/ {client_fio or "______________________"} /', SIG),
            Paragraph('Дата: ________________', SIG),
        ],
        [
            Paragraph('', SIG),
            Paragraph('(подпись)', HINT),
            Paragraph('(расшифровка подписи)', HINT),
            Paragraph('', SIG),
        ],
    ]

    sig_tbl = Table(sig_data, colWidths=[3.0*cm, 4.5*cm, 5.5*cm, 4.2*cm])
    sig_tbl.setStyle(TableStyle([
        ('FONTNAME',     (0,0), (-1,-1), fn),
        ('FONTSIZE',     (0,0), (-1,-1), 9),
        ('VALIGN',       (0,0), (-1,-1), 'BOTTOM'),
        ('TOPPADDING',   (0,0), (-1,-1), 2),
        ('BOTTOMPADDING',(0,0), (-1,-1), 2),
    ]))
    story.append(KeepTogether(sig_tbl))

    # ── Legal protection text ─────────────────────
    story.append(Spacer(1, 0.5*cm))
    story.append(Paragraph(
        'Подписывая настоящий акт, Клиент подтверждает, что выполненные работы приняты '
        'в полном объёме, в установленные сроки, претензий к объёму, качеству и стоимости '
        'выполненных работ и использованных материалов не имеется. '
        'Настоящий документ является основанием для передачи транспортного средства.',
        LEGAL
    ))

    doc.build(story)
    buf.seek(0)
    return buf


@login_required
def order_customer_pdf(request, pk):
    """Customer-facing work order PDF — preliminary or final based on order status."""
    order = get_object_or_404(Order, pk=pk)
    is_final = order.status in ('Готов', 'Завершён')
    buf = _make_customer_pdf(order, is_final=is_final)
    prefix = 'final' if is_final else 'preview'
    resp = HttpResponse(buf, content_type='application/pdf')
    resp['Content-Disposition'] = f'inline; filename="order_{order.pk}_{prefix}.pdf"'
    return resp


MODEL_VERBOSE = {
    'Client': 'Клиент',
    'ClientCar': 'Автомобиль',
    'Order': 'Заказ',
    'CarMake': 'Марка авто',
    'CarModel': 'Модель авто',
    'ServiceType': 'Тип услуги',
    'Service': 'Услуга',
    'Brand': 'Бренд',
    'Supplier': 'Поставщик',
    'Part': 'Запчасть',
    'StorageLocation': 'Локация',
    'PurchaseOrder': 'Закупка',
    'SupplyDocument': 'Поставка',
    'Employee': 'Сотрудник',
}


